"""Fase 9 — Importacion masiva de Compras Nacionalizadas desde Excel.

Alcance (ver CONTEXTO_CONTINUACION_IMPORTACION_COMPRAS.md): la mercaderia
YA esta nacionalizada al momento de importar. NO se modela FOB/CIF,
fletes, seguros ni aduanas -- esos campos (DUA, Factura, Pais origen,
Fecha documento) se guardan solo como dato de referencia/trazabilidad en
la orden de compra, tal como llega en el Excel.

Regla de oro de esta fase: NO se crea logica nueva de inventario. Este
modulo unicamente parsea el Excel, valida todo, y si es valido reutiliza
exactamente el flujo ya existente de m04_compras:

    crear_orden() -> aprobar_orden() -> recibir_orden()

recibir_orden() a su vez llama a inventario_service.registrar_ingreso()
(m03), sin cambios de logica ahi -- Fase 9 solo le pasa lote/fechas
reales del Excel en vez de dejar que se autogeneren (ver m04_compras/
service.py::recibir_orden, condicion `item.lote or f"OC-..."`).

Flujo en dos pasos, sin persistir una "carga" en base de datos (a
diferencia de m21_importacion_datos): previsualizar() NUNCA escribe;
confirmar() vuelve a validar el mismo archivo y recien ahi escribe. El
Frontend reenvia el mismo archivo seleccionado en ambos pasos (ver
pages/compras/compras.js).
"""
import io
import unicodedata
from datetime import datetime

import openpyxl
from sqlalchemy.orm import Session

from app.modules.m02_productos import repository as productos_repo
from app.modules.m03_inventario import service as inventario_service
from app.modules.m04_compras import importacion_schemas as schemas
from app.modules.m04_compras import schemas as compras_schemas
from app.modules.m04_compras import service as compras_service
from app.modules.m05_proveedores import repository as proveedores_repo
# Reutiliza el normalizador de moneda ya existente en m21 (Dolares/Soles ->
# USD/PEN) en vez de duplicar la logica aqui: el Excel real del cliente
# (COMPRAS_ECO_NEOAGROX_2026) trae la columna "Moneda" como texto libre
# ("Dolares"/"Soles"), no como codigo ISO de 3 letras.
from app.modules.m21_importacion_datos.service import _normalizar_moneda

COLUMNAS_OBLIGATORIAS = [
    "Orden Compra",
    "Proveedor",
    "Producto",
    "Cantidad",
    "Costo Unitario",
]

# "Lote" ya no es obligatoria: el formato Excel del cliente
# (COMPRAS_ECO_NEOAGROX_2026) no la trae. Si el Excel SI la trae, se sigue
# usando tal cual (y se sigue validando que no este duplicada); si no viene
# o la columna no existe, se autogenera (ver _validar_y_resolver_fila).

# Encabezado normalizado (sin tildes, minusculas, espacios colapsados) -> nombre canonico.
ALIAS_COLUMNAS = {
    "orden compra": "Orden Compra",
    "orden de compra": "Orden Compra",
    "pedido": "Orden Compra",  # formato Excel del cliente (COMPRAS_ECO_NEOAGROX_2026)
    "proveedor": "Proveedor",
    "producto": "Producto",
    "concepto": "Producto",  # formato del cliente: nombre del producto/insumo
    "descripcion": "Producto",  # idem, hoja 2 del Excel del cliente
    "cantidad": "Cantidad",
    "costo unitario": "Costo Unitario",
    "precio de compra": "Costo Unitario",  # formato del cliente
    "lote": "Lote",
    "fecha elaboracion": "Fecha Elaboracion",
    "fecha de elaboracion": "Fecha Elaboracion",
    "fecha vencimiento": "Fecha Vencimiento",
    "fecha de vencimiento": "Fecha Vencimiento",
    "factura": "Factura",
    "dua": "DUA",
    "pais origen": "Pais Origen",
    "pais de origen": "Pais Origen",
    "fecha documento": "Fecha Documento",
    "fecha del documento": "Fecha Documento",
    "fecha de emision factura": "Fecha Documento",  # formato del cliente
    "fecha de emision de factura": "Fecha Documento",
    "observaciones": "Observaciones",
    "comentario": "Observaciones",  # formato del cliente, hoja 2
    "comentarios": "Observaciones",
    # --- Adicionales, formato Excel del cliente (COMPRAS_ECO_NEOAGROX_2026).
    "moneda": "Moneda",
    "presentacion": "Presentacion",
    "unidad de medida": "Unidad Medida",
    "unida de medida": "Unidad Medida",  # tal como viene escrito en el Excel del cliente
    "cantidad por unidad": "Cantidad Por Unidad",
    "dias de credito": "Dias De Credito",
    "ruc": "RUC",
}

COLUMNAS_FECHA = {"Fecha Elaboracion", "Fecha Vencimiento", "Fecha Documento"}


def _sin_tildes(texto: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", texto) if unicodedata.category(c) != "Mn")


def _clave_encabezado(valor) -> str | None:
    if valor is None:
        return None
    texto = " ".join(str(valor).split())
    if not texto:
        return None
    return _sin_tildes(texto).lower()


def _normalizar_encabezado(valor) -> str | None:
    clave = _clave_encabezado(valor)
    if clave is None:
        return None
    return ALIAS_COLUMNAS.get(clave)


def _normalizar_fecha(valor):
    """Devuelve datetime o None. Acepta celdas de fecha reales de Excel y
    texto 'dd/mm/aaaa' o 'dd-mm-aaaa'. Si no se puede interpretar, devuelve
    el valor original (la validacion de la fila lo marcara como invalido)."""
    if valor in (None, ""):
        return None
    if isinstance(valor, datetime):
        return valor
    if hasattr(valor, "isoformat") and not isinstance(valor, str):
        return datetime(valor.year, valor.month, valor.day)
    texto = str(valor).strip()
    for sep in ("/", "-"):
        partes = texto.split(sep)
        if len(partes) == 3 and all(p.strip().isdigit() for p in partes):
            dia, mes, anio = (p.strip() for p in partes)
            anio = f"20{anio}" if len(anio) == 2 else anio
            try:
                return datetime(int(anio), int(mes), int(dia))
            except ValueError:
                pass
    return texto  # invalido; se detecta en la validacion de la fila


def _detectar_encabezados(hoja) -> tuple[int, dict[str, int]]:
    for num_fila, fila in enumerate(hoja.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        indices: dict[str, int] = {}
        for idx, celda in enumerate(fila):
            canonico = _normalizar_encabezado(celda)
            if canonico and canonico not in indices:
                indices[canonico] = idx
        if all(c in indices for c in COLUMNAS_OBLIGATORIAS):
            return num_fila, indices
    raise ValueError(
        "No se encontraron las columnas obligatorias en el Excel: " + ", ".join(COLUMNAS_OBLIGATORIAS)
    )


def _leer_filas_excel(contenido: bytes) -> tuple[list[dict], list[str]]:
    """Lee TODAS las hojas del libro (antes solo se leia wb.worksheets[0] y
    la(s) hoja(s) siguientes se ignoraban en silencio -- ver Excel real del
    cliente COMPRAS_ECO_NEOAGROX_2026, que trae 2 hojas). Cada hoja se
    procesa de forma independiente (encabezados propios, arrastre de
    'Orden Compra' combinada reiniciado por hoja); las hojas sin las
    columnas obligatorias se ignoran y se reportan en errores_por_hoja. Si
    NINGUNA hoja aporta filas, se lanza un error explicito con el detalle
    por hoja en vez de devolver una lista vacia sin explicacion. Devuelve
    (filas, hojas_procesadas) para que previsualizar()/confirmar() informen
    explicitamente que hojas se usaron."""
    wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
    filas: list[dict] = []
    hojas_procesadas: list[str] = []
    errores_por_hoja: dict[str, str] = {}

    for hoja in wb.worksheets:
        try:
            fila_encabezado, indices = _detectar_encabezados(hoja)
        except ValueError as exc:
            errores_por_hoja[hoja.title] = str(exc)
            continue
        hojas_procesadas.append(hoja.title)

        ultimo_orden_compra = None
        for num_fila, fila in enumerate(
            hoja.iter_rows(min_row=fila_encabezado + 1, values_only=True), start=fila_encabezado + 1
        ):
            valores = {}
            for col, idx in indices.items():
                valor = fila[idx] if idx < len(fila) else None
                valores[col] = _normalizar_fecha(valor) if col in COLUMNAS_FECHA else valor
            if not any(v not in (None, "") for v in valores.values()):
                continue  # fila vacia
            # El Excel del cliente agrupa los items de un mismo "Pedido" con la
            # celda de "Orden Compra" combinada verticalmente (una sola celda
            # visual para varias filas). openpyxl solo devuelve el valor en la
            # celda superior de la combinacion y None en las demas: se arrastra
            # hacia abajo el ultimo valor visto para no perder la asociacion
            # fila-pedido. El arrastre se reinicia en cada hoja nueva.
            if valores.get("Orden Compra") not in (None, ""):
                ultimo_orden_compra = valores["Orden Compra"]
            elif ultimo_orden_compra is not None:
                valores["Orden Compra"] = ultimo_orden_compra
            filas.append({"numero_fila": num_fila, "datos": valores, "hoja": hoja.title})

    if not filas and errores_por_hoja:
        detalle = "; ".join(f"'{h}': {m}" for h, m in errores_por_hoja.items())
        raise ValueError(
            "No se pudo leer ninguna fila del Excel: ninguna de sus hojas tiene las "
            f"columnas obligatorias {COLUMNAS_OBLIGATORIAS}. Detalle por hoja: {detalle}"
        )
    return filas, hojas_procesadas


def _texto(valor) -> str | None:
    if valor is None:
        return None
    texto = str(valor).strip()
    return texto or None


def _fecha_a_texto(valor) -> str | None:
    if isinstance(valor, datetime):
        return valor.date().isoformat()
    return None


def _validar_y_resolver_fila(db: Session, datos: dict, lotes_vistos_en_archivo: set[str], numero_fila: int = 0) -> tuple[str | None, dict]:
    """Devuelve (mensaje_error_o_None, resueltos). `resueltos` trae
    proveedor_id/producto_id cuando la fila es valida, para no tener que
    volver a buscarlos en confirmar()."""
    resueltos: dict = {}

    faltantes = [c for c in COLUMNAS_OBLIGATORIAS if not datos.get(c)]
    if faltantes:
        return f"Faltan columnas obligatorias: {faltantes}", resueltos

    proveedor = proveedores_repo.obtener_por_razon_social(db, str(datos["Proveedor"]).strip())
    if proveedor is None and datos.get("RUC"):
        # Formato Excel del cliente: si no matchea por razon social, se
        # intenta por RUC (columna adicional que no siempre trae el nombre
        # exactamente igual al catalogo de proveedores).
        proveedor = proveedores_repo.obtener_por_ruc(db, str(datos["RUC"]).strip())
    if proveedor is None:
        return f"Proveedor '{datos['Proveedor']}' no existe", resueltos
    if not proveedor.activo:
        return f"Proveedor '{datos['Proveedor']}' esta inactivo", resueltos
    resueltos["proveedor_id"] = proveedor.id

    producto = productos_repo.obtener_por_nombre(db, str(datos["Producto"]).strip())
    if producto is None:
        return f"Producto '{datos['Producto']}' no existe", resueltos
    resueltos["producto_id"] = producto.id
    resueltos["producto_perecible"] = producto.perecible

    try:
        cantidad = float(datos["Cantidad"])
    except (TypeError, ValueError):
        return "Cantidad no es un numero valido", resueltos
    if cantidad <= 0:
        return "Cantidad debe ser mayor a 0", resueltos
    resueltos["cantidad"] = cantidad

    try:
        costo = float(datos["Costo Unitario"])
    except (TypeError, ValueError):
        return "Costo Unitario no es un numero valido", resueltos
    if costo <= 0:
        return "Costo Unitario debe ser mayor a 0", resueltos
    resueltos["costo_unitario"] = costo

    lote = str(datos.get("Lote") or "").strip()
    if not lote:
        # Formato Excel del cliente no trae columna "Lote": se autogenera
        # igual que en el flujo manual (ver service.py::recibir_orden,
        # condicion `item.lote or f"OC-..."`), en vez de rechazar la fila.
        lote = f"{str(datos['Orden Compra']).strip()}-F{numero_fila}"
    clave_lote = lote.lower()
    if clave_lote in lotes_vistos_en_archivo:
        return f"Lote '{lote}' duplicado dentro del mismo archivo", resueltos
    if inventario_service.existe_lote_con_codigo(db, lote):
        return f"Lote '{lote}' ya existe en el inventario (no se puede duplicar)", resueltos
    resueltos["lote"] = lote

    fecha_elaboracion = datos.get("Fecha Elaboracion")
    if fecha_elaboracion is not None and not isinstance(fecha_elaboracion, datetime):
        return "Fecha Elaboracion no es una fecha valida (use dd/mm/aaaa)", resueltos
    resueltos["fecha_elaboracion"] = fecha_elaboracion

    fecha_vencimiento = datos.get("Fecha Vencimiento")
    if fecha_vencimiento is not None and not isinstance(fecha_vencimiento, datetime):
        return "Fecha Vencimiento no es una fecha valida (use dd/mm/aaaa)", resueltos
    if producto.perecible and fecha_vencimiento is None:
        return f"Producto '{producto.nombre}' es perecible: Fecha Vencimiento es obligatoria", resueltos
    resueltos["fecha_vencimiento"] = fecha_vencimiento

    fecha_documento = datos.get("Fecha Documento")
    if fecha_documento is not None and not isinstance(fecha_documento, datetime):
        return "Fecha Documento no es una fecha valida (use dd/mm/aaaa)", resueltos
    resueltos["fecha_documento"] = fecha_documento

    lotes_vistos_en_archivo.add(clave_lote)
    resueltos["orden_compra"] = str(datos["Orden Compra"]).strip()

    # --- Campos opcionales, formato Excel del cliente (COMPRAS_ECO_NEOAGROX_2026).
    # Ninguno es obligatorio: si faltan o no son validos, simplemente no se
    # guardan (no rechazan la fila), salvo Cantidad Por Unidad y Dias de
    # Credito que si vienen deben ser numericos.
    resueltos["concepto"] = str(datos["Producto"]).strip()
    # Bug corregido: el Excel real trae "Dolares"/"Soles" (texto) y el
    # schema de OrdenCompra exige un codigo ISO de 3 letras (USD/PEN). Antes
    # este valor se pasaba tal cual, y confirmar() fallaba con un error
    # crudo de Pydantic pese a que previsualizar() reportaba 0 errores.
    moneda_original = _texto(datos.get("Moneda"))
    resueltos["moneda"] = _normalizar_moneda(moneda_original, None) if moneda_original else None
    resueltos["presentacion"] = _texto(datos.get("Presentacion"))
    resueltos["unidad_medida"] = _texto(datos.get("Unidad Medida"))

    cantidad_por_unidad = datos.get("Cantidad Por Unidad")
    if cantidad_por_unidad not in (None, ""):
        if not _es_numero(cantidad_por_unidad):
            return "Cantidad Por Unidad no es un numero valido", resueltos
        resueltos["cantidad_por_unidad"] = float(cantidad_por_unidad)
    else:
        resueltos["cantidad_por_unidad"] = None

    dias_credito = datos.get("Dias De Credito")
    if dias_credito not in (None, ""):
        if not _es_numero(dias_credito):
            return "Dias de Credito no es un numero valido", resueltos
        resueltos["dias_credito"] = int(dias_credito)
    else:
        resueltos["dias_credito"] = None

    return None, resueltos


def _fila_a_out(
    numero_fila: int, datos: dict, valida: bool, mensaje_error: str | None, hoja: str | None = None
) -> schemas.FilaImportacionCompraOut:
    return schemas.FilaImportacionCompraOut(
        numero_fila=numero_fila,
        hoja=hoja,
        orden_compra=_texto(datos.get("Orden Compra")),
        proveedor=_texto(datos.get("Proveedor")),
        producto=_texto(datos.get("Producto")),
        cantidad=float(datos["Cantidad"]) if _es_numero(datos.get("Cantidad")) else None,
        costo_unitario=float(datos["Costo Unitario"]) if _es_numero(datos.get("Costo Unitario")) else None,
        lote=_texto(datos.get("Lote")),
        fecha_elaboracion=_fecha_a_texto(datos.get("Fecha Elaboracion")),
        fecha_vencimiento=_fecha_a_texto(datos.get("Fecha Vencimiento")),
        factura=_texto(datos.get("Factura")),
        dua=_texto(datos.get("DUA")),
        pais_origen=_texto(datos.get("Pais Origen")),
        fecha_documento=_fecha_a_texto(datos.get("Fecha Documento")),
        observaciones=_texto(datos.get("Observaciones")),
        moneda=_texto(datos.get("Moneda")),
        presentacion=_texto(datos.get("Presentacion")),
        unidad_medida=_texto(datos.get("Unidad Medida")),
        cantidad_por_unidad=float(datos["Cantidad Por Unidad"]) if _es_numero(datos.get("Cantidad Por Unidad")) else None,
        dias_credito=int(float(datos["Dias De Credito"])) if _es_numero(datos.get("Dias De Credito")) else None,
        ruc=_texto(datos.get("RUC")),
        valida=valida,
        mensaje_error=mensaje_error,
    )


def _es_numero(valor) -> bool:
    try:
        float(valor)
        return True
    except (TypeError, ValueError):
        return False


def previsualizar(db: Session, inventario_destino_id: int, nombre_archivo: str, contenido: bytes) -> schemas.PreviewImportacionComprasOut:
    if not nombre_archivo.lower().endswith((".xlsx", ".xls")):
        raise ValueError("El archivo debe ser un Excel (.xlsx o .xls)")

    inventario_service.obtener_inventario(db, inventario_destino_id)  # valida existencia

    filas_excel, hojas_procesadas = _leer_filas_excel(contenido)
    if not filas_excel:
        raise ValueError("El archivo no tiene filas con datos")

    lotes_vistos: set[str] = set()
    filas_out: list[schemas.FilaImportacionCompraOut] = []
    ordenes_externas: set[str] = set()
    validas = 0
    for f in filas_excel:
        mensaje_error, resueltos = _validar_y_resolver_fila(db, f["datos"], lotes_vistos, f["numero_fila"])
        es_valida = mensaje_error is None
        if es_valida:
            validas += 1
            ordenes_externas.add(resueltos["orden_compra"])
        filas_out.append(_fila_a_out(f["numero_fila"], f["datos"], es_valida, mensaje_error, f.get("hoja")))

    return schemas.PreviewImportacionComprasOut(
        nombre_archivo=nombre_archivo,
        inventario_destino_id=inventario_destino_id,
        total_filas=len(filas_excel),
        filas_validas=validas,
        filas_con_error=len(filas_excel) - validas,
        ordenes_a_crear=len(ordenes_externas),
        hojas_procesadas=hojas_procesadas,
        filas=filas_out,
    )


def confirmar(db: Session, inventario_destino_id: int, nombre_archivo: str, contenido: bytes) -> schemas.ConfirmarImportacionComprasOut:
    """Vuelve a validar todo el archivo (nada se da por bueno de una
    previsualizacion anterior) y, solo con lo valido, ejecuta
    crear_orden -> aprobar_orden -> recibir_orden agrupando filas por
    'Orden Compra'. Las filas invalidas se omiten y se reportan; no
    detienen la importacion de las ordenes que si son validas."""
    if not nombre_archivo.lower().endswith((".xlsx", ".xls")):
        raise ValueError("El archivo debe ser un Excel (.xlsx o .xls)")

    inventario_service.obtener_inventario(db, inventario_destino_id)  # valida existencia

    filas_excel, hojas_procesadas = _leer_filas_excel(contenido)
    if not filas_excel:
        raise ValueError("El archivo no tiene filas con datos")

    lotes_vistos: set[str] = set()
    grupos: dict[str, list[dict]] = {}
    fallidas: list[schemas.FilaImportacionCompraOut] = []

    for f in filas_excel:
        mensaje_error, resueltos = _validar_y_resolver_fila(db, f["datos"], lotes_vistos, f["numero_fila"])
        if mensaje_error is not None:
            fallidas.append(_fila_a_out(f["numero_fila"], f["datos"], False, mensaje_error, f.get("hoja")))
            continue
        clave_orden = resueltos["orden_compra"]
        grupos.setdefault(clave_orden, []).append({
            "numero_fila": f["numero_fila"],
            "datos": f["datos"],
            "resueltos": resueltos,
        })

    ordenes_creadas: list[schemas.OrdenCreadaImportacionOut] = []
    filas_procesadas = 0

    for numero_orden_externo, filas_grupo in grupos.items():
        proveedores_del_grupo = {f["resueltos"]["proveedor_id"] for f in filas_grupo}
        if len(proveedores_del_grupo) > 1:
            for f in filas_grupo:
                fallidas.append(_fila_a_out(
                    f["numero_fila"], f["datos"], False,
                    f"La Orden Compra '{numero_orden_externo}' trae mas de un Proveedor distinto en el archivo",
                ))
            continue

        primera = filas_grupo[0]["datos"]
        try:
            orden = compras_service.crear_orden(
                db,
                compras_schemas.OrdenCompraCrear(
                    proveedor_id=filas_grupo[0]["resueltos"]["proveedor_id"],
                    inventario_destino_id=inventario_destino_id,
                    moneda=filas_grupo[0]["resueltos"].get("moneda") or "USD",
                    observaciones=_texto(primera.get("Observaciones")) or (
                        f"Importacion de compras — Pedido/Orden Compra {numero_orden_externo}"
                    ),
                    numero_orden_externo=numero_orden_externo,
                    invoice=_texto(primera.get("Factura")),
                    documento_aduanero=_texto(primera.get("DUA")),
                    pais_origen=_texto(primera.get("Pais Origen")),
                    fecha_documento=primera.get("Fecha Documento") if isinstance(primera.get("Fecha Documento"), datetime) else None,
                    dias_credito=filas_grupo[0]["resueltos"].get("dias_credito"),
                    items=[
                        compras_schemas.OrdenCompraItemCrear(
                            producto_id=f["resueltos"]["producto_id"],
                            cantidad=f["resueltos"]["cantidad"],
                            costo_unitario=f["resueltos"]["costo_unitario"],
                            lote=f["resueltos"]["lote"],
                            fecha_elaboracion=f["resueltos"]["fecha_elaboracion"],
                            fecha_vencimiento=f["resueltos"]["fecha_vencimiento"],
                            observaciones=_texto(f["datos"].get("Observaciones")),
                            presentacion=f["resueltos"].get("presentacion"),
                            unidad_medida=f["resueltos"].get("unidad_medida"),
                            cantidad_por_unidad=f["resueltos"].get("cantidad_por_unidad"),
                            concepto=f["resueltos"].get("concepto"),
                        )
                        for f in filas_grupo
                    ],
                ),
            )
            orden = compras_service.aprobar_orden(db, orden.id)
            orden = compras_service.recibir_orden(db, orden.id)
        except Exception as exc:
            db.rollback()
            for f in filas_grupo:
                fallidas.append(_fila_a_out(f["numero_fila"], f["datos"], False, str(exc)))
            continue

        filas_procesadas += len(filas_grupo)
        ordenes_creadas.append(schemas.OrdenCreadaImportacionOut(
            numero_orden_externo=numero_orden_externo,
            orden_compra_id=orden.id,
            estado=orden.estado,
            items_creados=len(filas_grupo),
        ))

    return schemas.ConfirmarImportacionComprasOut(
        nombre_archivo=nombre_archivo,
        inventario_destino_id=inventario_destino_id,
        ordenes_creadas=ordenes_creadas,
        filas_procesadas=filas_procesadas,
        filas_fallidas=fallidas,
        hojas_procesadas=hojas_procesadas,
    )
