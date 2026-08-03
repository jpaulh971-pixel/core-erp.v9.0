"""Exportadores de reportes a Excel (.xlsx) y PDF, para m19_reportes.

No consulta la base de datos directamente: recibe los mismos objetos
Pydantic que ya devuelven los endpoints JSON (schemas.ReporteVentas,
ReporteCompras, ReporteInventarioValorizado, ResumenGeneral) y los
convierte a bytes descargables. Asi se evita duplicar ninguna consulta
de `repository.py` / logica de `service.py`.
"""
from __future__ import annotations

from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.modules.m19_reportes import schemas

TITULO_FONT = Font(bold=True, size=13)
ENCABEZADO_FONT = Font(bold=True)


def _rango_texto(desde: date | None, hasta: date | None) -> str:
    if desde is None and hasta is None:
        return "Todo el periodo"
    return f"{desde or '(inicio)'}  →  {hasta or '(hoy)'}"


# ---------------------------------------------------------------------
# Excel: helper generico (una hoja por seccion, headers + filas)
# ---------------------------------------------------------------------
def _excel_desde_hojas(titulo: str, subtitulo: str, hojas: list[tuple[str, list[str], list[list]]]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)  # se crea una hoja por seccion, ninguna por defecto

    for nombre_hoja, headers, filas in hojas:
        ws = wb.create_sheet(title=nombre_hoja[:31])  # Excel limita a 31 caracteres
        ws["A1"] = titulo
        ws["A1"].font = TITULO_FONT
        ws["A2"] = subtitulo
        ws.append([])  # fila 3 en blanco
        ws.append(headers)
        for celda in ws[4]:
            celda.font = ENCABEZADO_FONT
        for fila in filas:
            ws.append(fila)
        for columna in ws.columns:
            largo = max((len(str(c.value)) for c in columna if c.value is not None), default=10)
            ws.column_dimensions[columna[0].column_letter].width = min(largo + 2, 45)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------
# PDF: helper generico (una tabla por seccion, con su propio titulo)
# ---------------------------------------------------------------------
def _pdf_desde_secciones(titulo: str, subtitulo: str, secciones: list[tuple[str, list[str], list[list]]]) -> bytes:
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm, leftMargin=1.5 * cm, rightMargin=1.5 * cm,
    )
    estilos = getSampleStyleSheet()
    elementos = [
        Paragraph(titulo, estilos["Title"]),
        Paragraph(subtitulo, estilos["Normal"]),
        Spacer(1, 0.5 * cm),
    ]

    for nombre_seccion, headers, filas in secciones:
        elementos.append(Paragraph(nombre_seccion, estilos["Heading2"]))
        datos_tabla = [headers] + [[str(v) for v in fila] for fila in filas] if filas else [headers, ["(sin datos)"] * len(headers)]
        tabla = Table(datos_tabla, repeatRows=1)
        tabla.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f3f4f6")]),
                ]
            )
        )
        elementos.append(tabla)
        elementos.append(Spacer(1, 0.6 * cm))

    doc.build(elementos)
    return buffer.getvalue()


# ---------------------------------------------------------------------
# Ventas
# ---------------------------------------------------------------------
def _costo_unitario_texto(valor: float | None) -> str:
    return f"{valor:.4f}" if valor is not None else "—"


def _filas_ventas(reporte: schemas.ReporteVentas) -> tuple[list, list]:
    por_producto = []
    for p in reporte.por_producto:
        # costo_total = costo_unitario_promedio * cantidad: multiplicacion
        # de dos campos ya presentes (derivado, no un calculo de costeo
        # nuevo). "Total" (venta) se mantiene aparte porque es ingreso, no
        # costo -- no deben confundirse en la misma columna.
        costo_total = (
            p.costo_unitario_promedio * p.cantidad
            if p.costo_unitario_promedio is not None
            else None
        )
        por_producto.append(
            [
                p.codigo, p.nombre, p.cantidad,
                _costo_unitario_texto(p.costo_unitario_promedio),
                _costo_unitario_texto(costo_total),
                round(p.total, 2),
            ]
        )
    por_cliente = [[c.razon_social, c.cantidad_ordenes, round(c.total, 2)] for c in reporte.por_cliente]
    return por_producto, por_cliente


def ventas_excel(reporte: schemas.ReporteVentas) -> bytes:
    por_producto, por_cliente = _filas_ventas(reporte)
    subtitulo = f"{_rango_texto(reporte.desde, reporte.hasta)}  |  {reporte.total_ordenes} orden(es)  |  Total vendido: {reporte.total_vendido:.2f}"
    return _excel_desde_hojas(
        "Reporte de Ventas", subtitulo,
        [
            (
                "Por producto",
                ["Codigo", "Producto", "Cantidad", "Costo Unitario", "Costo Total", "Total Vendido"],
                por_producto,
            ),
            ("Por cliente", ["Cliente", "Cant. ordenes", "Total"], por_cliente),
        ],
    )


def ventas_pdf(reporte: schemas.ReporteVentas) -> bytes:
    por_producto, por_cliente = _filas_ventas(reporte)
    subtitulo = f"{_rango_texto(reporte.desde, reporte.hasta)} — {reporte.total_ordenes} orden(es) — Total vendido: {reporte.total_vendido:.2f}"
    return _pdf_desde_secciones(
        "Reporte de Ventas", subtitulo,
        [
            (
                "Por producto",
                ["Codigo", "Producto", "Cantidad", "Costo Unitario", "Costo Total", "Total Vendido"],
                por_producto,
            ),
            ("Por cliente", ["Cliente", "Cant. ordenes", "Total"], por_cliente),
        ],
    )


# ---------------------------------------------------------------------
# Compras
# ---------------------------------------------------------------------
def _filas_compras(reporte: schemas.ReporteCompras) -> tuple[list, list]:
    por_producto = []
    for p in reporte.por_producto:
        # costo_total = costo_unitario_promedio * cantidad: multiplicacion
        # de dos campos ya presentes (derivado, no un calculo de costeo
        # nuevo), mismo patron ya aplicado en _filas_ventas. "Total Compra"
        # es el monto ya sumado por el repository desde OrdenCompraItem.
        costo_total = (
            p.costo_unitario_promedio * p.cantidad
            if p.costo_unitario_promedio is not None
            else None
        )
        por_producto.append(
            [
                p.codigo, p.nombre, p.cantidad,
                _costo_unitario_texto(p.costo_unitario_promedio),
                _costo_unitario_texto(costo_total),
                round(p.total, 2),
            ]
        )
    por_proveedor = [[p.razon_social, p.cantidad_ordenes, round(p.total, 2)] for p in reporte.por_proveedor]
    return por_producto, por_proveedor


def compras_excel(reporte: schemas.ReporteCompras) -> bytes:
    por_producto, por_proveedor = _filas_compras(reporte)
    subtitulo = f"{_rango_texto(reporte.desde, reporte.hasta)}  |  {reporte.total_ordenes} orden(es)  |  Total comprado: {reporte.total_comprado:.2f}"
    return _excel_desde_hojas(
        "Reporte de Compras", subtitulo,
        [
            (
                "Por producto",
                ["Codigo", "Producto", "Cantidad", "Costo Unitario", "Costo Total", "Total Compra"],
                por_producto,
            ),
            ("Por proveedor", ["Proveedor", "Cant. ordenes", "Total"], por_proveedor),
        ],
    )


def compras_pdf(reporte: schemas.ReporteCompras) -> bytes:
    por_producto, por_proveedor = _filas_compras(reporte)
    subtitulo = f"{_rango_texto(reporte.desde, reporte.hasta)} — {reporte.total_ordenes} orden(es) — Total comprado: {reporte.total_comprado:.2f}"
    return _pdf_desde_secciones(
        "Reporte de Compras", subtitulo,
        [
            (
                "Por producto",
                ["Codigo", "Producto", "Cantidad", "Costo Unitario", "Costo Total", "Total Compra"],
                por_producto,
            ),
            ("Por proveedor", ["Proveedor", "Cant. ordenes", "Total"], por_proveedor),
        ],
    )


# ---------------------------------------------------------------------
# Inventario valorizado
# ---------------------------------------------------------------------
def _filas_inventario(reporte: schemas.ReporteInventarioValorizado) -> list:
    return [
        [
            p.codigo, p.nombre, p.cantidad_actual, round(p.valor_promedio_unitario, 4),
            round(p.valor_total, 2), p.stock_minimo, "SI" if p.bajo_stock_minimo else "NO",
        ]
        for p in reporte.productos
    ]


def inventario_valorizado_excel(reporte: schemas.ReporteInventarioValorizado) -> bytes:
    subtitulo = (
        f"Generado: {reporte.generado_en:%Y-%m-%d %H:%M}  |  {reporte.total_productos} producto(s)  |  "
        f"Valor total: {reporte.valor_total_inventario:.2f}  |  Bajo stock minimo: {reporte.productos_bajo_stock_minimo}"
    )
    headers = ["Codigo", "Producto", "Cantidad", "Vr. unit. prom.", "Valor total", "Stock minimo", "Bajo minimo"]
    return _excel_desde_hojas(
        "Inventario Valorizado", subtitulo,
        [("Inventario", headers, _filas_inventario(reporte))],
    )


def inventario_valorizado_pdf(reporte: schemas.ReporteInventarioValorizado) -> bytes:
    subtitulo = (
        f"Generado: {reporte.generado_en:%Y-%m-%d %H:%M} — {reporte.total_productos} producto(s) — "
        f"Valor total: {reporte.valor_total_inventario:.2f} — Bajo stock minimo: {reporte.productos_bajo_stock_minimo}"
    )
    headers = ["Codigo", "Producto", "Cant.", "Vr.unit.", "Vr.total", "St.min", "Bajo min"]
    return _pdf_desde_secciones(
        "Inventario Valorizado", subtitulo,
        [("Inventario", headers, _filas_inventario(reporte))],
    )


# ---------------------------------------------------------------------
# Resumen general
# ---------------------------------------------------------------------
def _filas_resumen(reporte: schemas.ResumenGeneral) -> list:
    return [
        ["Total vendido (periodo)", round(reporte.total_vendido_periodo, 2)],
        ["Total comprado (periodo)", round(reporte.total_comprado_periodo, 2)],
        ["Ordenes de venta (periodo)", reporte.ordenes_venta_periodo],
        ["Ordenes de compra (periodo)", reporte.ordenes_compra_periodo],
        ["Valor de inventario actual", round(reporte.valor_inventario_actual, 2)],
        ["Productos bajo stock minimo", reporte.productos_bajo_stock_minimo],
    ]


def resumen_general_excel(reporte: schemas.ResumenGeneral) -> bytes:
    subtitulo = _rango_texto(reporte.desde, reporte.hasta)
    return _excel_desde_hojas(
        "Resumen General", subtitulo,
        [("Resumen", ["Indicador", "Valor"], _filas_resumen(reporte))],
    )


def resumen_general_pdf(reporte: schemas.ResumenGeneral) -> bytes:
    subtitulo = _rango_texto(reporte.desde, reporte.hasta)
    return _pdf_desde_secciones(
        "Resumen General", subtitulo,
        [("Resumen", ["Indicador", "Valor"], _filas_resumen(reporte))],
    )


# ---------------------------------------------------------------------
# FASE 2 - control gerencial para inventario perecible.
# Mismo patron que el resto del archivo: no consultan la base de datos,
# reciben los mismos schemas Pydantic que ya devuelven los endpoints
# JSON (schemas.ReporteInventarioPorLote, schemas.ReporteProximosVencer)
# y los convierten a bytes descargables via los mismos helpers
# genericos (_excel_desde_hojas / _pdf_desde_secciones) ya usados por
# ventas/compras/inventario-valorizado/resumen-general.
# ---------------------------------------------------------------------
def _fecha_texto(valor: datetime | None) -> str:
    return f"{valor:%Y-%m-%d}" if valor is not None else "—"


def _dias_restantes_texto(valor: int | None) -> str:
    return str(valor) if valor is not None else "—"


def _filas_inventario_por_lote(reporte: schemas.ReporteInventarioPorLote) -> list:
    return [
        [
            l.codigo_producto, l.producto, l.codigo_lote,
            _fecha_texto(l.fecha_vencimiento), l.cantidad_disponible,
            round(l.costo_unitario, 4), round(l.valor_total_lote, 2),
            l.estado_lote, l.semaforo_vencimiento,
            _dias_restantes_texto(l.dias_restantes_vencimiento),
            l.proveedor or "—",
        ]
        for l in reporte.lotes
    ]


def inventario_por_lote_excel(reporte: schemas.ReporteInventarioPorLote) -> bytes:
    subtitulo = (
        f"Generado: {reporte.generado_en:%Y-%m-%d %H:%M}  |  {reporte.total_lotes} lote(s)  |  "
        f"Valor total: {reporte.valor_total:.2f}"
    )
    headers = [
        "Codigo", "Producto", "Lote", "Vencimiento", "Cant. disponible",
        "Costo Unitario", "Valor Total", "Estado", "Semaforo", "Dias restantes", "Proveedor",
    ]
    return _excel_desde_hojas(
        "Inventario por Lote", subtitulo,
        [("Lotes", headers, _filas_inventario_por_lote(reporte))],
    )


def inventario_por_lote_pdf(reporte: schemas.ReporteInventarioPorLote) -> bytes:
    subtitulo = (
        f"Generado: {reporte.generado_en:%Y-%m-%d %H:%M} — {reporte.total_lotes} lote(s) — "
        f"Valor total: {reporte.valor_total:.2f}"
    )
    headers = [
        "Codigo", "Producto", "Lote", "Vencim.", "Cant.disp.",
        "Costo Unit.", "Vr.Total", "Estado", "Semaforo", "Dias", "Proveedor",
    ]
    return _pdf_desde_secciones(
        "Inventario por Lote", subtitulo,
        [("Lotes", headers, _filas_inventario_por_lote(reporte))],
    )


def _filas_proximos_vencer(reporte: schemas.ReporteProximosVencer) -> list:
    return [
        [
            l.codigo_producto, l.producto, l.codigo_lote,
            _fecha_texto(l.fecha_vencimiento), _dias_restantes_texto(l.dias_restantes),
            l.categoria, l.estado_lote, l.semaforo_vencimiento,
            l.cantidad_disponible, round(l.costo_unitario, 4),
            round(l.valor_stock_comprometido, 2),
        ]
        for l in reporte.lotes
    ]


def proximos_vencer_excel(reporte: schemas.ReporteProximosVencer) -> bytes:
    subtitulo = (
        f"Generado: {reporte.generado_en:%Y-%m-%d %H:%M}  |  {reporte.total_lotes} lote(s)  |  "
        f"Activos: {reporte.activos}  |  Proximos a vencer: {reporte.proximos_a_vencer}  |  "
        f"Vencidos: {reporte.vencidos}  |  Valor comprometido: {reporte.valor_total_comprometido:.2f}"
    )
    headers = [
        "Codigo", "Producto", "Lote", "Vencimiento", "Dias restantes",
        "Categoria", "Estado", "Semaforo", "Cant. disponible", "Costo Unitario", "Valor comprometido",
    ]
    return _excel_desde_hojas(
        "Proximos a Vencer", subtitulo,
        [("Lotes", headers, _filas_proximos_vencer(reporte))],
    )


def proximos_vencer_pdf(reporte: schemas.ReporteProximosVencer) -> bytes:
    subtitulo = (
        f"Generado: {reporte.generado_en:%Y-%m-%d %H:%M} — {reporte.total_lotes} lote(s) — "
        f"Activos: {reporte.activos} — Proximos a vencer: {reporte.proximos_a_vencer} — "
        f"Vencidos: {reporte.vencidos} — Valor comprometido: {reporte.valor_total_comprometido:.2f}"
    )
    headers = [
        "Codigo", "Producto", "Lote", "Vencim.", "Dias",
        "Categoria", "Estado", "Semaforo", "Cant.disp.", "Costo Unit.", "Vr.comprometido",
    ]
    return _pdf_desde_secciones(
        "Proximos a Vencer", subtitulo,
        [("Lotes", headers, _filas_proximos_vencer(reporte))],
    )
