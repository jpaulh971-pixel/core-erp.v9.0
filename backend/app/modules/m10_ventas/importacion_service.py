"""Fase 10 — Importacion masiva de Ventas desde Excel.

Regla de oro de esta fase (identica a la Fase 9 de Compras): NO se crea
logica nueva de inventario, Kardex ni FEFO. Este modulo unicamente
parsea el Excel, valida todo, y si es valido reutiliza exactamente el
flujo ya existente de m10_ventas:

    crear_orden() -> confirmar_orden() -> despachar_orden()

despachar_orden() a su vez llama a inventario_service.registrar_salida()
(m03) SIN pasar lote_id: el consumo de stock sigue siendo FEFO
automatico, exactamente igual que una venta manual. La columna opcional
"Lote" del Excel del cliente se guarda unicamente como referencia en la
fila de previsualizacion; no selecciona ni fuerza ningun lote concreto
del Kardex.

Diferencia deliberada frente a Compras (Fase 9): esta importacion es
TODO-O-NADA. Compras crea las ordenes validas y reporta las filas
invalidas aparte (continua). Ventas, en cambio, si encuentra una sola
fila invalida en confirmar() no crea absolutamente nada -- lanza
ValueError (-> 400) y el cliente debe corregir el Excel completo antes
de reintentar. Esto incluye una validacion de stock agregada por
producto+inventario ANTES de escribir nada, precisamente para que un
archivo que paso la validacion de filas no pueda fallar a mitad de
camino por falta de stock.

Flujo en dos pasos, sin persistir una "carga" en base de datos (igual
que Compras): previsualizar() NUNCA escribe; confirmar() vuelve a
validar el mismo archivo y recien ahi escribe.
"""
import io
import unicodedata
from datetime import datetime

import openpyxl
from sqlalchemy.orm import Session

from app.modules.m02_productos import repository as productos_repo
from app.modules.m03_inventario import repository as inventario_repo
from app.modules.m03_inventario import service as inventario_service
from app.modules.m09_moneda import validators as moneda_validators
# Reutiliza el mismo normalizador de moneda que Compras (Dolares/Soles ->
# USD/PEN), en vez de duplicar la logica aqui: el Excel real del cliente
# (VENTAS_2026_-_INVENTARIO_1) trae la columna "Moneda" como texto libre.
from app.modules.m21_importacion_datos.service import _normalizar_moneda
from app.modules.m10_ventas import importacion_schemas as schemas
from app.modules.m10_ventas import repository
from app.modules.m10_ventas import schemas as ventas_schemas
from app.modules.m10_ventas import service as ventas_service
from app.modules.m11_clientes import repository as clientes_repo

COLUMNAS_OBLIGATORIAS = [
    "Orden de Venta",
    "Moneda",
    "Cantidad",
    "Descripcion",
    "Precio Venta",
    "RUC",
]

# Encabezado normalizado (sin tildes, minusculas, espacios colapsados) -> nombre canonico.
ALIAS_COLUMNAS = {
    "orden de compra/venta": "Orden de Venta",
    "orden compra/venta": "Orden de Venta",
    "orden de venta": "Orden de Venta",
    "orden venta": "Orden de Venta",
    # Formato Excel real del cliente (VENTAS_2026_-_INVENTARIO_1): la
    # columna se llama literalmente "Orden de Compra" aunque es el numero
    # de pedido del cliente para esta venta (a diferencia de Compras, que
    # usa "Pedido" para lo mismo del lado del proveedor).
    "orden de compra": "Orden de Venta",
    "orden compra": "Orden de Venta",
    "vendedor": "Vendedor",
    "moneda": "Moneda",
    "cantidad": "Cantidad",
    "unidad de medida": "Unidad de Medida",
    "unidad medida": "Unidad de Medida",
    "descripcion": "Descripcion",
    "producto": "Descripcion",
    "codigo del producto": "Codigo Producto",
    "codigo de producto": "Codigo Producto",
    "codigo producto": "Codigo Producto",
    "precio venta": "Precio Venta",
    "precio de venta": "Precio Venta",
    "subtotal": "Subtotal",
    "sub total": "Subtotal",
    "igv": "IGV",
    "total": "Total",
    "fecha de emision": "Fecha Emision",
    "fecha emision": "Fecha Emision",
    "dias de credito": "Dias Credito",
    "dias credito": "Dias Credito",
    "fecha de vencimiento": "Fecha Vencimiento",
    "fecha vencimiento": "Fecha Vencimiento",
    "factura": "Factura",
    "estado": "Estado",
    "ruc": "RUC",
    "cliente": "Cliente",
    "ano": "Anio",  # "Año" sin tilde
    "anio": "Anio",
    "mes": "Mes",
    "guia de remision": "Guia Remision",
    "guia remision": "Guia Remision",
    "cultivo": "Cultivo",
    "fundo": "Fundo",
    "lote": "Lote",
    "observaciones": "Observaciones",
}

COLUMNAS_FECHA = {"Fecha Emision", "Fecha Vencimiento"}


def _sin_tildes(texto: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", texto) if unicodedata.category(c) != "Mn")


def _clave_encabezado(valor) -> str | None:
    if valor is None:
        return None
    texto = " ".join(str(valor).split())
    if not texto:
        return None
    return _sin_tildes(texto).lower()


def _normalizar_encabezado(valor) -> str | None:
    clave = _clave_encabezado(valor)
    if clave is None:
        return None
    return ALIAS_COLUMNAS.get(clave)


def _normalizar_fecha(valor):
    """Devuelve datetime o None. Acepta celdas de fecha reales de Excel y
    texto 'dd/mm/aaaa' o 'dd-mm-aaaa'. Si no se puede interpretar, devuelve
    el valor original (la validacion de la fila lo marcara como invalido)."""
    if valor in (None, ""):
        return None
    if isinstance(valor, datetime):
        return valor
    if hasattr(valor, "isoformat") and not isinstance(valor, str):
        return datetime(valor.year, valor.month, valor.day)
    texto = str(valor).strip()
    for sep in ("/", "-"):
        partes = texto.split(sep)
        if len(partes) == 3 and all(p.strip().isdigit() for p in partes):
            dia, mes, anio = (p.strip() for p in partes)
            anio = f"20{anio}" if len(anio) == 2 else anio
            try:
                return datetime(int(anio), int(mes), int(dia))
            except ValueError:
                pass
    return texto  # invalido; se detecta en la validacion de la fila


def _detectar_encabezados(hoja) -> tuple[int, dict[str, int]]:
    for num_fila, fila in enumerate(hoja.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        indices: dict[str, int] = {}
        for idx, celda in enumerate(fila):
            canonico = _normalizar_encabezado(celda)
            if canonico and canonico not in indices:
                indices[canonico] = idx
        if all(c in indices for c in COLUMNAS_OBLIGATORIAS):
            return num_fila, indices
    raise ValueError(
        "No se encontraron las columnas obligatorias en el Excel: " + ", ".join(COLUMNAS_OBLIGATORIAS)
    )


def _leer_filas_excel(contenido: bytes) -> tuple[list[dict], list[str]]:
    """Lee TODAS las hojas del libro (antes solo se leia wb.worksheets[0],
    igual que el bug de Compras: cualquier hoja adicional se ignoraba en
    silencio). Cada hoja se procesa de forma independiente (encabezados
    propios); las hojas sin las columnas obligatorias se ignoran y se
    reportan en errores_por_hoja. Si NINGUNA hoja aporta filas, se lanza un
    error explicito con el detalle por hoja. Devuelve (filas,
    hojas_procesadas) para que previsualizar()/confirmar() informen
    explicitamente que hojas se usaron."""
    wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
    filas: list[dict] = []
    hojas_procesadas: list[str] = []
    errores_por_hoja: dict[str, str] = {}

    for hoja in wb.worksheets:
        try:
            fila_encabezado, indices = _detectar_encabezados(hoja)
        except ValueError as exc:
            errores_por_hoja[hoja.title] = str(exc)
            continue
        hojas_procesadas.append(hoja.title)

        for num_fila, fila in enumerate(
            hoja.iter_rows(min_row=fila_encabezado + 1, values_only=True), start=fila_encabezado + 1
        ):
            valores = {}
            for col, idx in indices.items():
                valor = fila[idx] if idx < len(fila) else None
                valores[col] = _normalizar_fecha(valor) if col in COLUMNAS_FECHA else valor
            if not any(v not in (None, "") for v in valores.values()):
                continue  # fila vacia
            filas.append({"numero_fila": num_fila, "datos": valores, "hoja": hoja.title})

    if not filas and errores_por_hoja:
        detalle = "; ".join(f"'{h}': {m}" for h, m in errores_por_hoja.items())
        raise ValueError(
            "No se pudo leer ninguna fila del Excel: ninguna de sus hojas tiene las "
            f"columnas obligatorias {COLUMNAS_OBLIGATORIAS}. Detalle por hoja: {detalle}"
        )
    return filas, hojas_procesadas


def _texto(valor) -> str | None:
    if valor is None:
        return None
    texto = str(valor).strip()
    return texto or None


def _fecha_a_texto(valor) -> str | None:
    if isinstance(valor, datetime):
        return valor.date().isoformat()
    return None


def _es_numero(valor) -> bool:
    try:
        float(valor)
        return True
    except (TypeError, ValueError):
        return False


def _es_entero(valor) -> bool:
    try:
        int(float(valor))
        return True
    except (TypeError, ValueError):
        return False


def _validar_moneda(moneda: str) -> bool:
    """Mismo criterio que m09_moneda.validators.validar_codigo_moneda,
    pero como chequeo booleano (para no propagar HTTPException fila por
    fila): codigo ISO-4217 de 3 letras."""
    return len(moneda) == 3 and moneda.isalpha()


def _validar_y_resolver_fila(db: Session, datos: dict) -> tuple[str | None, dict]:
    """Devuelve (mensaje_error_o_None, resueltos). `resueltos` trae
    cliente_id/producto_id ya buscados cuando la fila es valida, para no
    tener que volver a buscarlos en confirmar()."""
    resueltos: dict = {}

    faltantes = [c for c in COLUMNAS_OBLIGATORIAS if not datos.get(c)]
    if faltantes:
        return f"Faltan columnas obligatorias: {faltantes}", resueltos

    ruc = str(datos["RUC"]).strip()
    cliente = clientes_repo.obtener_por_ruc(db, ruc)
    if cliente is None:
        return f"No existe un cliente con RUC '{ruc}'", resueltos
    if not cliente.activo:
        return f"El cliente con RUC '{ruc}' esta inactivo", resueltos
    resueltos["cliente_id"] = cliente.id
    resueltos["ruc_cliente"] = ruc

    codigo_producto = _texto(datos.get("Codigo Producto"))
    producto = None
    if codigo_producto:
        producto = productos_repo.obtener_por_codigo(db, codigo_producto)
    if producto is None:
        producto = productos_repo.obtener_por_nombre(db, str(datos["Descripcion"]).strip())
    if producto is None:
        referencia = codigo_producto or str(datos["Descripcion"]).strip()
        return f"Producto '{referencia}' no existe", resueltos
    resueltos["producto_id"] = producto.id

    # Bug corregido: el Excel real trae "Dolares"/"Soles" (texto libre) en
    # la columna Moneda, y la validacion exigia exactamente un codigo
    # ISO-4217 de 3 letras, por lo que TODAS las filas fallaban. Se
    # normaliza primero (mismo normalizador que Compras) y solo se
    # rechaza si, tras normalizar, sigue sin ser un codigo ISO valido.
    moneda_normalizada = _normalizar_moneda(datos["Moneda"], "")
    if not _validar_moneda(moneda_normalizada):
        return (
            f"Moneda '{datos['Moneda']}' invalida. Debe ser ISO-4217 de 3 letras o un valor "
            "reconocido (ej: PEN, USD, Soles, Dolares)",
            resueltos,
        )
    resueltos["moneda"] = moneda_normalizada

    try:
        cantidad = float(datos["Cantidad"])
    except (TypeError, ValueError):
        return "Cantidad no es un numero valido", resueltos
    if cantidad <= 0:
        return "Cantidad debe ser mayor a 0", resueltos
    resueltos["cantidad"] = cantidad

    try:
        precio_venta = float(datos["Precio Venta"])
    except (TypeError, ValueError):
        return "Precio Venta no es un numero valido", resueltos
    if precio_venta < 0:
        return "Precio Venta no puede ser negativo", resueltos
    resueltos["precio_venta"] = precio_venta

    for campo_opcional_numerico in ("Subtotal", "IGV", "Total"):
        valor = datos.get(campo_opcional_numerico)
        if valor not in (None, "") and not _es_numero(valor):
            return f"{campo_opcional_numerico} no es un numero valido", resueltos

    dias_credito = datos.get("Dias Credito")
    if dias_credito not in (None, "") and not _es_entero(dias_credito):
        return "Dias Credito no es un numero entero valido", resueltos

    anio = datos.get("Anio")
    if anio not in (None, "") and not _es_entero(anio):
        return "Anio no es un numero entero valido", resueltos

    fecha_emision = datos.get("Fecha Emision")
    if fecha_emision is not None and not isinstance(fecha_emision, datetime):
        return "Fecha Emision no es una fecha valida (use dd/mm/aaaa)", resueltos

    fecha_vencimiento = datos.get("Fecha Vencimiento")
    if fecha_vencimiento is not None and not isinstance(fecha_vencimiento, datetime):
        return "Fecha Vencimiento no es una fecha valida (use dd/mm/aaaa)", resueltos

    # Nota: la unicidad de Factura/Guia de Remision DENTRO del archivo se
    # valida a nivel de grupo (una misma "Orden de Venta" -> una misma
    # Factura en todas sus filas es normal: son lineas del mismo
    # comprobante). Aqui solo se valida contra ordenes YA EXISTENTES en
    # la base de datos, que es un chequeo valido fila por fila.
    factura = _texto(datos.get("Factura"))
    if factura and repository.existe_factura(db, factura):
        return f"Factura '{factura}' ya existe en Ventas (no se puede duplicar)", resueltos

    guia_remision = _texto(datos.get("Guia Remision"))
    if guia_remision and repository.existe_guia_remision(db, guia_remision):
        return f"Guia de Remision '{guia_remision}' ya existe en Ventas (no se puede duplicar)", resueltos

    resueltos["orden_venta"] = str(datos["Orden de Venta"]).strip()
    return None, resueltos


def _fila_a_out(
    numero_fila: int, datos: dict, valida: bool, mensaje_error: str | None, hoja: str | None = None
) -> schemas.FilaImportacionVentaOut:
    return schemas.FilaImportacionVentaOut(
        numero_fila=numero_fila,
        hoja=hoja,
        orden_venta=_texto(datos.get("Orden de Venta")),
        vendedor=_texto(datos.get("Vendedor")),
        moneda=_texto(datos.get("Moneda")),
        cantidad=float(datos["Cantidad"]) if _es_numero(datos.get("Cantidad")) else None,
        unidad_medida=_texto(datos.get("Unidad de Medida")),
        descripcion=_texto(datos.get("Descripcion")),
        codigo_producto=_texto(datos.get("Codigo Producto")),
        precio_venta=float(datos["Precio Venta"]) if _es_numero(datos.get("Precio Venta")) else None,
        sub_total=float(datos["Subtotal"]) if _es_numero(datos.get("Subtotal")) else None,
        igv=float(datos["IGV"]) if _es_numero(datos.get("IGV")) else None,
        total=float(datos["Total"]) if _es_numero(datos.get("Total")) else None,
        fecha_emision=_fecha_a_texto(datos.get("Fecha Emision")),
        dias_credito=int(float(datos["Dias Credito"])) if _es_entero(datos.get("Dias Credito")) else None,
        fecha_vencimiento=_fecha_a_texto(datos.get("Fecha Vencimiento")),
        factura=_texto(datos.get("Factura")),
        estado=_texto(datos.get("Estado")),
        ruc=_texto(datos.get("RUC")),
        cliente=_texto(datos.get("Cliente")),
        anio=int(float(datos["Anio"])) if _es_entero(datos.get("Anio")) else None,
        mes=_texto(datos.get("Mes")),
        guia_remision=_texto(datos.get("Guia Remision")),
        cultivo=_texto(datos.get("Cultivo")),
        fundo=_texto(datos.get("Fundo")),
        lote=_texto(datos.get("Lote")),
        observaciones=_texto(datos.get("Observaciones")),
        valida=valida,
        mensaje_error=mensaje_error,
    )


def _validar_todas_las_filas(
    db: Session, filas_excel: list[dict], inventario_salida_id: int
) -> tuple[list[schemas.FilaImportacionVentaOut], list[dict], int]:
    """Valida fila por fila y, adicionalmente, valida el STOCK AGREGADO
    por producto (sumando la cantidad de todas las filas validas de ese
    producto) contra el inventario de salida indicado -- solo lectura,
    sin tocar Kardex/FEFO. Devuelve (filas_out, filas_validas_con_resueltos, cantidad_validas).
    """
    filas_out: list[schemas.FilaImportacionVentaOut] = []
    validas: list[dict] = []

    for f in filas_excel:
        mensaje_error, resueltos = _validar_y_resolver_fila(db, f["datos"])
        es_valida = mensaje_error is None
        if es_valida:
            validas.append({"numero_fila": f["numero_fila"], "datos": f["datos"], "resueltos": resueltos})
        filas_out.append(_fila_a_out(f["numero_fila"], f["datos"], es_valida, mensaje_error, f.get("hoja")))

    # --- Validacion de stock agregado (solo lectura), producto por producto.
    demanda_por_producto: dict[int, float] = {}
    for fila in validas:
        pid = fila["resueltos"]["producto_id"]
        demanda_por_producto[pid] = demanda_por_producto.get(pid, 0.0) + fila["resueltos"]["cantidad"]

    productos_sin_stock: dict[int, str] = {}
    for producto_id, cantidad_requerida in demanda_por_producto.items():
        producto_inventario = inventario_repo.obtener_producto_inventario(db, producto_id, inventario_salida_id)
        stock_disponible = (
            inventario_repo.stock_total_producto_inventario(db, producto_inventario.id)
            if producto_inventario is not None
            else 0.0
        )
        if stock_disponible < cantidad_requerida:
            productos_sin_stock[producto_id] = (
                f"Stock insuficiente en el inventario seleccionado: se requieren "
                f"{cantidad_requerida} unidades en total y solo hay {stock_disponible} disponibles"
            )

    if productos_sin_stock:
        # Reescribe filas_out y validas: cualquier fila cuyo producto no
        # tiene stock agregado suficiente pasa a invalida.
        nuevas_filas_out: list[schemas.FilaImportacionVentaOut] = []
        nuevas_validas: list[dict] = []
        for fila_out, fila_original in zip(filas_out, filas_excel):
            if not fila_out.valida:
                nuevas_filas_out.append(fila_out)
                continue
            fila_valida_data = next(
                (v for v in validas if v["numero_fila"] == fila_out.numero_fila), None
            )
            producto_id = fila_valida_data["resueltos"]["producto_id"] if fila_valida_data else None
            if producto_id in productos_sin_stock:
                nuevas_filas_out.append(
                    _fila_a_out(
                        fila_out.numero_fila,
                        fila_original["datos"],
                        False,
                        productos_sin_stock[producto_id],
                        fila_original.get("hoja"),
                    )
                )
            else:
                nuevas_filas_out.append(fila_out)
                if fila_valida_data:
                    nuevas_validas.append(fila_valida_data)
        filas_out = nuevas_filas_out
        validas = nuevas_validas

    return filas_out, validas, len(validas)


def previsualizar(db: Session, inventario_salida_id: int, nombre_archivo: str, contenido: bytes) -> schemas.PreviewImportacionVentasOut:
    if not nombre_archivo.lower().endswith((".xlsx", ".xls")):
        raise ValueError("El archivo debe ser un Excel (.xlsx o .xls)")

    inventario_service.obtener_inventario(db, inventario_salida_id)  # valida existencia

    filas_excel, hojas_procesadas = _leer_filas_excel(contenido)
    if not filas_excel:
        raise ValueError("El archivo no tiene filas con datos")

    filas_out, validas, cantidad_validas = _validar_todas_las_filas(db, filas_excel, inventario_salida_id)
    ordenes_externas = {v["resueltos"]["orden_venta"] for v in validas}

    return schemas.PreviewImportacionVentasOut(
        nombre_archivo=nombre_archivo,
        inventario_salida_id=inventario_salida_id,
        total_filas=len(filas_excel),
        filas_validas=cantidad_validas,
        filas_con_error=len(filas_excel) - cantidad_validas,
        ordenes_a_crear=len(ordenes_externas),
        hojas_procesadas=hojas_procesadas,
        filas=filas_out,
    )


def confirmar(db: Session, inventario_salida_id: int, nombre_archivo: str, contenido: bytes) -> schemas.ConfirmarImportacionVentasOut:
    """Vuelve a validar todo el archivo (nada se da por bueno de una
    previsualizacion anterior). A diferencia de Compras, esta importacion
    es TODO-O-NADA: si queda una sola fila invalida (incluyendo la
    validacion de stock agregado), no se crea NADA y se lanza ValueError
    (-> 400 en el router) con el detalle de las filas con error.

    Solo si el archivo completo es valido, agrupa por 'Orden de Venta' y
    ejecuta crear_orden -> confirmar_orden -> despachar_orden para cada
    grupo, reutilizando exactamente el flujo manual existente.
    """
    if not nombre_archivo.lower().endswith((".xlsx", ".xls")):
        raise ValueError("El archivo debe ser un Excel (.xlsx o .xls)")

    inventario_service.obtener_inventario(db, inventario_salida_id)  # valida existencia

    filas_excel, hojas_procesadas = _leer_filas_excel(contenido)
    if not filas_excel:
        raise ValueError("El archivo no tiene filas con datos")

    filas_out, validas, cantidad_validas = _validar_todas_las_filas(db, filas_excel, inventario_salida_id)

    filas_invalidas = [f for f in filas_out if not f.valida]
    if filas_invalidas:
        detalle = "; ".join(
            f"fila {f.numero_fila}: {f.mensaje_error}" for f in filas_invalidas
        )
        raise ValueError(
            f"El archivo tiene {len(filas_invalidas)} fila(s) invalida(s); no se creo ninguna "
            f"orden de venta (importacion todo-o-nada). Detalle: {detalle}"
        )

    grupos: dict[str, list[dict]] = {}
    for v in validas:
        grupos.setdefault(v["resueltos"]["orden_venta"], []).append(v)

    # --- Consistencia por grupo: mismo cliente y misma moneda dentro de
    # una misma "Orden de Venta" (igual criterio que Compras exige mismo
    # proveedor por "Orden Compra").
    for numero_orden_externo, filas_grupo in grupos.items():
        clientes_del_grupo = {f["resueltos"]["cliente_id"] for f in filas_grupo}
        monedas_del_grupo = {f["resueltos"]["moneda"] for f in filas_grupo}
        if len(clientes_del_grupo) > 1:
            raise ValueError(
                f"La Orden de Venta '{numero_orden_externo}' trae mas de un Cliente (RUC) distinto en el archivo"
            )
        if len(monedas_del_grupo) > 1:
            raise ValueError(
                f"La Orden de Venta '{numero_orden_externo}' trae mas de una Moneda distinta en el archivo"
            )

    # --- Unicidad de Factura/Guia de Remision ENTRE distintas Ordenes de
    # Venta del mismo archivo. Dentro de un mismo grupo, todas las filas
    # comparten la misma Factura/Guia (son lineas del mismo comprobante):
    # eso es normal y no es un error.
    facturas_por_grupo: dict[str, str] = {}
    guias_por_grupo: dict[str, str] = {}
    for numero_orden_externo, filas_grupo in grupos.items():
        factura_grupo = _texto(filas_grupo[0]["datos"].get("Factura"))
        if factura_grupo:
            clave = factura_grupo.lower()
            if clave in facturas_por_grupo and facturas_por_grupo[clave] != numero_orden_externo:
                raise ValueError(
                    f"La Factura '{factura_grupo}' aparece en mas de una Orden de Venta distinta "
                    f"('{facturas_por_grupo[clave]}' y '{numero_orden_externo}') dentro del archivo"
                )
            facturas_por_grupo[clave] = numero_orden_externo

        guia_grupo = _texto(filas_grupo[0]["datos"].get("Guia Remision"))
        if guia_grupo:
            clave = guia_grupo.lower()
            if clave in guias_por_grupo and guias_por_grupo[clave] != numero_orden_externo:
                raise ValueError(
                    f"La Guia de Remision '{guia_grupo}' aparece en mas de una Orden de Venta distinta "
                    f"('{guias_por_grupo[clave]}' y '{numero_orden_externo}') dentro del archivo"
                )
            guias_por_grupo[clave] = numero_orden_externo

    ordenes_creadas: list[schemas.OrdenCreadaImportacionVentaOut] = []
    ordenes_ya_creadas_ids: list[int] = []
    filas_procesadas = 0

    try:
        for numero_orden_externo, filas_grupo in grupos.items():
            primera = filas_grupo[0]["datos"]
            primera_resuelta = filas_grupo[0]["resueltos"]

            orden = ventas_service.crear_orden(
                db,
                ventas_schemas.OrdenVentaCrear(
                    cliente_id=primera_resuelta["cliente_id"],
                    inventario_salida_id=inventario_salida_id,
                    moneda=primera_resuelta["moneda"],
                    observaciones=_texto(primera.get("Observaciones"))
                    or f"Importacion de ventas — Orden de Venta {numero_orden_externo}",
                    numero_orden_externo=numero_orden_externo,
                    vendedor=_texto(primera.get("Vendedor")),
                    factura=_texto(primera.get("Factura")),
                    guia_remision=_texto(primera.get("Guia Remision")),
                    fecha_emision=primera.get("Fecha Emision") if isinstance(primera.get("Fecha Emision"), datetime) else None,
                    dias_credito=int(float(primera["Dias Credito"])) if _es_entero(primera.get("Dias Credito")) else None,
                    fecha_vencimiento=primera.get("Fecha Vencimiento") if isinstance(primera.get("Fecha Vencimiento"), datetime) else None,
                    estado_documento=_texto(primera.get("Estado")),
                    ruc_cliente=primera_resuelta["ruc_cliente"],
                    anio=int(float(primera["Anio"])) if _es_entero(primera.get("Anio")) else None,
                    meses=_texto(primera.get("Mes")),
                    cultivo=_texto(primera.get("Cultivo")),
                    fundo=_texto(primera.get("Fundo")),
                    items=[
                        ventas_schemas.OrdenVentaItemCrear(
                            producto_id=f["resueltos"]["producto_id"],
                            cantidad=f["resueltos"]["cantidad"],
                            precio_unitario_venta=f["resueltos"]["precio_venta"],
                            unidad_medida=_texto(f["datos"].get("Unidad de Medida")),
                            descripcion=_texto(f["datos"].get("Descripcion")),
                            sub_total=float(f["datos"]["Subtotal"]) if _es_numero(f["datos"].get("Subtotal")) else None,
                            igv=float(f["datos"]["IGV"]) if _es_numero(f["datos"].get("IGV")) else None,
                            total=float(f["datos"]["Total"]) if _es_numero(f["datos"].get("Total")) else None,
                        )
                        for f in filas_grupo
                    ],
                ),
            )
            ordenes_ya_creadas_ids.append(orden.id)
            orden = ventas_service.confirmar_orden(db, orden.id)
            orden = ventas_service.despachar_orden(db, orden.id)

            filas_procesadas += len(filas_grupo)
            ordenes_creadas.append(
                schemas.OrdenCreadaImportacionVentaOut(
                    numero_orden_externo=numero_orden_externo,
                    orden_venta_id=orden.id,
                    estado=orden.estado,
                    items_creados=len(filas_grupo),
                )
            )
    except Exception as exc:
        # Todo-o-nada: si un grupo falla (caso excepcional, ya que el
        # stock agregado y las filas se validaron por completo antes de
        # escribir nada), se intenta cancelar -- SOLO estado, sin tocar
        # Kardex/FEFO -- las ordenes de este mismo confirmar() que aun
        # esten en un estado cancelable. Las que ya quedaron DESPACHADA
        # no se pueden revertir por la maquina de estados existente (no
        # se modifica esa regla): quedan reportadas en el detalle del
        # error para que el usuario las revise manualmente.
        no_revertidas = []
        for orden_id in ordenes_ya_creadas_ids:
            orden_existente = ventas_service.obtener_orden(db, orden_id)
            if orden_existente.estado in ("BORRADOR", "CONFIRMADA"):
                ventas_service.cancelar_orden(db, orden_id)
            else:
                no_revertidas.append(orden_id)
        detalle_no_revertidas = (
            f" Ordenes ya despachadas antes de la falla (no se revirtieron): {no_revertidas}."
            if no_revertidas
            else ""
        )
        raise ValueError(
            f"Fallo al crear/confirmar/despachar la importacion: {exc}.{detalle_no_revertidas}"
        )

    return schemas.ConfirmarImportacionVentasOut(
        nombre_archivo=nombre_archivo,
        inventario_salida_id=inventario_salida_id,
        ordenes_creadas=ordenes_creadas,
        filas_procesadas=filas_procesadas,
        hojas_procesadas=hojas_procesadas,
    )
