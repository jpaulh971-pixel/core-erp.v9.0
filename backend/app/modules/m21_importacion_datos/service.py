"""Carga de Inventario Inicial desde el Excel del cliente.

Mapa Excel -> ERP (confirmado con el cliente para el Excel "Inventario
inicial 31/12/2025"):

    Producto              -> producto.nombre (get-or-create en m02)
    Familia                -> producto_inventario.familia
    Presentacion            -> producto_inventario.presentacion
    Litros Presentacion    -> producto_inventario.litros_presentacion
    Codigo interno          -> producto_inventario.codigo_interno
    Cantidad                -> lote.cantidad_inicial / cantidad_actual
    Lote                     -> lote.codigo_lote
    Fecha elaboracion       -> lote.fecha_elaboracion
    Fecha vencimiento       -> lote.fecha_vencimiento

Familia, Presentacion, Litros Presentacion y Codigo interno son
columnas opcionales en el Excel: si el cliente no las trae, se autogenera
un codigo_interno y el resto queda vacio (se puede completar despues a
mano en m02/m03).

Proceso obligatorio de dos pasos: previsualizar() NUNCA escribe en
m02/m03; confirmar() reutiliza inventario_service.registrar_ingreso,
que crea el ProductoInventario (si no existe) + Lote + Kardex de una.
"""
import io
import json
import time
import unicodedata
from contextlib import contextmanager
from datetime import datetime, timezone

import openpyxl
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.modules.m02_productos import repository as productos_repo
from app.modules.m02_productos.models import Producto
from app.modules.m03_inventario import schemas as inv_schemas
from app.modules.m03_inventario import service as inventario_service
from app.modules.m03_inventario.models import Lote, MovimientoKardex
from app.modules.m04_compras import repository as compras_repo
from app.modules.m04_compras import schemas as compras_schemas
from app.modules.m04_compras import service as compras_service
from app.modules.m04_compras.models import OrdenCompra
from app.modules.m05_proveedores import repository as proveedores_repo
from app.modules.m05_proveedores.models import Proveedor
from app.modules.m07_operacion_logistica.models import OperacionLogistica
from app.modules.m08_costos.models import CostoAdicional
from app.modules.m10_ventas import repository as ventas_repo
from app.modules.m10_ventas import schemas as ventas_schemas
from app.modules.m10_ventas import service as ventas_service
from app.modules.m10_ventas.models import OrdenVenta
from app.modules.m11_clientes import repository as clientes_repo
from app.modules.m11_clientes.models import Cliente
from app.modules.m17_guias_remision.models import GuiaRemision
from app.modules.m21_importacion_datos import repository, schemas, validators
from app.modules.m21_importacion_datos.models import (
    BitacoraReemplazo,
    CargaComprasHistorico,
    CargaComprasHistoricoFila,
    CargaInventarioInicial,
    CargaInventarioInicialFila,
    CargaVentasHistorico,
    CargaVentasHistoricoFila,
    ConfiguracionCorteInventario,
)

# Columnas obligatorias para reconocer la fila de encabezados y validar cada fila.
# "Costo unitario" es obligatoria (fix 2026-08-02): antes no se pedia y el
# costo del lote se grababa fijo en 0, lo que hacia que el valor de
# inventario (dashboard y reportes valorizados) saliera siempre en 0.00
# para todo lo cargado por este flujo.
COLUMNAS_OBLIGATORIAS = ["Producto", "Cantidad", "Lote", "Costo unitario"]

# Alias aceptados -> nombre canonico usado internamente (todo comparado sin
# tildes y en minusculas, para tolerar variaciones del Excel del cliente).
ALIAS_COLUMNAS = {
    "producto": "Producto",
    "familia": "Familia",
    "presentacion": "Presentacion",
    "litros presentacion": "Litros Presentacion",
    "codigo interno": "Codigo interno",
    "cantidad": "Cantidad",
    "lote": "Lote",
    "fecha elaboracion": "Fecha elaboracion",
    "fecha de elaboracion": "Fecha elaboracion",  # header real del cliente: "FECHA DE ELABORACION"
    "fecha vencimiento": "Fecha vencimiento",
    "fecha de vencimiento": "Fecha vencimiento",  # header real del cliente: "FECHA DE VENCIMIENTO"
    "costo unitario": "Costo unitario",
    "costo": "Costo unitario",
    "precio unitario": "Costo unitario",
    "precio de compra": "Costo unitario",
    "precio compra": "Costo unitario",
    # El Excel real de Saldo Inicial (INVENTARIO_PIURA_-_NEOAGROX) no trae
    # "Costo unitario", trae "Cantidad" y "Total": el costo unitario se
    # deriva como Total / Cantidad (ver _leer_filas_excel).
    "total": "Total",
}


def _sin_tildes(texto: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", texto) if unicodedata.category(c) != "Mn"
    )


def _clave_encabezado(valor) -> str | None:
    """Normaliza el texto de un encabezado para compararlo contra los
    alias: sin tildes, en minusculas y con saltos de linea/espacios
    multiples colapsados a uno solo (el Excel del cliente trae headers
    como 'Fecha de\nEmision Factura' partidos en dos lineas dentro de la
    misma celda)."""
    if valor is None:
        return None
    texto = " ".join(str(valor).split())
    if not texto:
        return None
    return _sin_tildes(texto).lower()


def _normalizar_encabezado(valor) -> str | None:
    clave = _clave_encabezado(valor)
    if clave is None:
        return None
    return ALIAS_COLUMNAS.get(clave)


def _normalizar_valor(valor):
    if isinstance(valor, datetime):
        return valor.isoformat()
    return valor


# --- Meses en espanol abreviados, para fechas de texto tipo '05-Ene-26' u
#     'Oct-25' (Excel del cliente trae columnas de fecha con ese formato
#     de despliegue; si la celda es texto y no fecha real, hay que
#     parsearla a mano). ---
_MESES_ES = {
    "ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
    "jul": 7, "ago": 8, "set": 9, "sep": 9, "oct": 10, "nov": 11, "dic": 12,
}


def _normalizar_fecha(valor):
    """Devuelve la fecha en formato ISO 'AAAA-MM-DD'. Acepta:
    - datetime/date reales (lo que entrega openpyxl para celdas con
      formato de fecha, sea cual sea el formato de despliegue: 'Oct-25',
      '05-Ene-26', etc. son solo formato visual, el valor subyacente ya
      es una fecha real y este caso es el mas comun);
    - texto 'dd/mm/aaaa' o 'dd-mm-aaaa';
    - texto 'aaaammdd' (8 digitos);
    - texto 'dd-Mon-aa' con mes en espanol abreviado (ej. '05-Ene-26');
    - texto 'Mon-aa' solo mes/anio (ej. 'Oct-25'): SUPUESTO documentado,
      se asume dia 1 del mes ante la ausencia de dia en el dato.
    Si el texto no calza con ningun formato reconocido, se devuelve el
    valor original sin modificar para que la validacion existente de la
    fila lo siga marcando como fecha invalida (mismo comportamiento que
    antes de este cambio)."""
    if valor in (None, ""):
        return valor
    if isinstance(valor, datetime):
        return valor.date().isoformat()
    if hasattr(valor, "isoformat") and not isinstance(valor, str):
        return valor.isoformat()  # datetime.date

    texto = str(valor).strip()

    try:
        datetime.fromisoformat(texto)
        return texto
    except ValueError:
        pass

    if texto.isdigit() and len(texto) == 8:
        try:
            return datetime.strptime(texto, "%Y%m%d").date().isoformat()
        except ValueError:
            pass

    for sep in ("/", "-"):
        partes = texto.split(sep)
        if len(partes) == 3 and all(p.strip().isdigit() for p in partes):
            dia, mes, anio = (p.strip() for p in partes)
            anio = f"20{anio}" if len(anio) == 2 else anio
            try:
                return datetime(int(anio), int(mes), int(dia)).date().isoformat()
            except ValueError:
                pass

    for sep in ("-", "/", " "):
        partes = texto.split(sep)
        if len(partes) == 3:
            dia_txt, mes_txt, anio_txt = (p.strip() for p in partes)
            mes_num = _MESES_ES.get(_sin_tildes(mes_txt).lower()[:3])
            if dia_txt.isdigit() and mes_num and anio_txt.isdigit():
                anio_txt = f"20{anio_txt}" if len(anio_txt) == 2 else anio_txt
                try:
                    return datetime(int(anio_txt), mes_num, int(dia_txt)).date().isoformat()
                except ValueError:
                    pass

    for sep in ("-", "/"):
        partes = texto.split(sep)
        if len(partes) == 2:
            mes_txt, anio_txt = (p.strip() for p in partes)
            mes_num = _MESES_ES.get(_sin_tildes(mes_txt).lower()[:3])
            if mes_num and anio_txt.isdigit():
                anio_txt = f"20{anio_txt}" if len(anio_txt) == 2 else anio_txt
                try:
                    return datetime(int(anio_txt), mes_num, 1).date().isoformat()
                except ValueError:
                    pass

    return valor


# --- Monedas: el Excel del cliente trae 'Dólares'/'Dolares'/'Soles' en
#     lugar del codigo ISO de 3 letras que espera el schema (Field con
#     min_length=3, max_length=3). Se normaliza a PEN/USD/EUR; si no se
#     reconoce el texto, se conserva el comportamiento anterior (primeras
#     3 letras en mayuscula) para no romper archivos que ya traian el
#     codigo ISO correcto. ---
_MONEDAS_EQUIVALENTES = {
    "pen": "PEN", "sol": "PEN", "soles": "PEN", "s/": "PEN", "s/.": "PEN",
    "usd": "USD", "dolar": "USD", "dolares": "USD", "us$": "USD",
    "dollar": "USD", "dollars": "USD",
    "eur": "EUR", "euro": "EUR", "euros": "EUR",
}


def _normalizar_moneda(valor, por_defecto: str) -> str:
    if not valor:
        return por_defecto
    clave = _sin_tildes(str(valor).strip()).lower()
    if clave in _MONEDAS_EQUIVALENTES:
        return _MONEDAS_EQUIVALENTES[clave]
    # Match por prefijo: el Excel real del cliente trae variantes como
    # 'Dolare' (columna angosta truncando "Dolares" en pantalla, pero a
    # veces tambien en el dato) o plurales/singulares mixtos.
    if clave.startswith("dolar") or clave.startswith("usd"):
        return "USD"
    if clave.startswith("sol") or clave == "pen":
        return "PEN"
    if clave.startswith("euro") or clave == "eur":
        return "EUR"
    return str(valor).strip()[:3].upper() or por_defecto


def _expandir_celdas_combinadas(hoja) -> None:
    """Los Excel de Saldo Inicial / Compras / Ventas del cliente combinan
    celdas verticalmente (ej. el nombre del producto o la presentacion
    combinados sobre varias filas de lote, ver captura 'Inventario
    Piura'). openpyxl solo deja el valor en la celda superior-izquierda
    del rango combinado; el resto queda en None. Se replica ese valor en
    todas las celdas del rango antes de leer fila por fila, para no
    perder el dato en las filas de abajo del combinado."""
    for rango in list(hoja.merged_cells.ranges):
        valor = hoja.cell(row=rango.min_row, column=rango.min_col).value
        # Las celdas no-superior-izquierda de un rango combinado son
        # MergedCell (solo lectura); hay que desmarcar el rango como
        # combinado antes de poder escribirles el valor propagado.
        coordenadas = str(rango)
        hoja.unmerge_cells(coordenadas)
        if valor is None:
            continue
        for fila in range(rango.min_row, rango.max_row + 1):
            for columna in range(rango.min_col, rango.max_col + 1):
                hoja.cell(row=fila, column=columna).value = valor


def _detectar_encabezados(hoja) -> tuple[int, dict[str, int]]:
    """Busca la fila de encabezados: la primera fila que contenga todas
    las columnas obligatorias. Devuelve (num_fila, {columna_canonica: indice})."""
    # El Excel real de Saldo Inicial no siempre trae "Costo unitario": si
    # trae "Cantidad" + "Total" se acepta igual, porque el costo unitario
    # se puede derivar (Total / Cantidad, ver _leer_filas_excel).
    obligatorias_sin_costo = [c for c in COLUMNAS_OBLIGATORIAS if c != "Costo unitario"]
    for i, fila in enumerate(hoja.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        indices: dict[str, int] = {}
        for idx, valor in enumerate(fila):
            canonica = _normalizar_encabezado(valor)
            if canonica:
                indices[canonica] = idx
        if all(c in indices for c in COLUMNAS_OBLIGATORIAS):
            return i, indices
        if all(c in indices for c in obligatorias_sin_costo) and "Total" in indices:
            return i, indices
    raise ValueError(
        f"No se encontraron los encabezados obligatorios {COLUMNAS_OBLIGATORIAS} "
        f"en las primeras 15 filas de la hoja '{hoja.title}'."
    )


# Columnas de Inventario Inicial que contienen fechas: requieren el
# parseo robusto de _normalizar_fecha (texto 'dd-Mon-aa', 'Mon-aa', etc.,
# ademas de fechas reales de Excel), a diferencia del resto de columnas
# que solo necesitan _normalizar_valor.
_COLUMNAS_FECHA_INVENTARIO = {"Fecha elaboracion", "Fecha vencimiento"}


def _leer_filas_excel(contenido: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
    filas_resultado = []
    errores_por_hoja: dict[str, str] = {}
    for hoja in wb.worksheets:
        try:
            fila_encabezado, indices = _detectar_encabezados(hoja)
        except ValueError as exc:
            # Hoja sin encabezados utiles (ej. hoja vacia o de notas): se
            # guarda el motivo para reportarlo si NINGUNA hoja del libro
            # resulta valida (ver chequeo despues del bucle). Si al menos
            # otra hoja si tiene encabezados validos, esta se ignora en
            # silencio como hasta ahora.
            errores_por_hoja[hoja.title] = str(exc)
            continue
        _expandir_celdas_combinadas(hoja)  # soporte Saldo Inicial con celdas combinadas
        for num_fila, fila in enumerate(
            hoja.iter_rows(min_row=fila_encabezado + 1, values_only=True), start=fila_encabezado + 1
        ):
            valores = {}
            for col, idx in indices.items():
                if idx >= len(fila):
                    continue
                valor = fila[idx]
                valores[col] = (
                    _normalizar_fecha(valor) if col in _COLUMNAS_FECHA_INVENTARIO else _normalizar_valor(valor)
                )
            if not any(v not in (None, "") for v in valores.values()):
                continue  # fila totalmente vacia
            # El Excel real de Saldo Inicial (INVENTARIO_PIURA_-_NEOAGROX)
            # no trae "Costo unitario", trae "Cantidad" y "Total": se
            # deriva el costo unitario como Total / Cantidad para no
            # obligar al cliente a agregar una columna que no tiene.
            if valores.get("Costo unitario") in (None, "") and valores.get("Total") not in (None, ""):
                try:
                    cantidad_num = float(valores.get("Cantidad"))
                    total_num = float(valores["Total"])
                    if cantidad_num != 0:
                        valores["Costo unitario"] = total_num / cantidad_num
                except (TypeError, ValueError):
                    pass  # se deja sin calcular; _validar_fila lo reportara como faltante/invalido
            valores["_hoja"] = hoja.title
            filas_resultado.append({"numero_fila": num_fila, "datos": valores})

    if not filas_resultado and errores_por_hoja:
        # Bug corregido: antes, si TODAS las hojas fallaban la deteccion de
        # encabezados, la funcion devolvia [] en silencio y el preview
        # terminaba en "0 filas" sin ninguna explicacion. Ahora, en ese
        # caso (ninguna hoja aporto una sola fila), se relanza un error
        # explicito con el detalle de cada hoja, para que el usuario sepa
        # exactamente que encabezado falto y en que hoja.
        detalle = "; ".join(f"'{h}': {m}" for h, m in errores_por_hoja.items())
        raise ValueError(
            f"No se pudo leer ninguna fila del Excel: ninguna de sus hojas tiene "
            f"los encabezados obligatorios {COLUMNAS_OBLIGATORIAS}. Detalle por hoja: {detalle}"
        )
    return filas_resultado


def _validar_fila(datos: dict) -> str | None:
    """Devuelve el mensaje de error si la fila es invalida, o None si es valida."""
    # "Costo unitario" en 0 es un valor valido (ej. muestras gratuitas), por
    # eso se valida aparte con su propio try/except en vez de con el chequeo
    # generico "not datos.get(c)" de abajo (que trataria 0 como faltante).
    faltantes = [c for c in COLUMNAS_OBLIGATORIAS if c != "Costo unitario" and not datos.get(c)]
    if "Costo unitario" not in datos or datos.get("Costo unitario") in (None, ""):
        faltantes.append("Costo unitario")
    if faltantes:
        return f"Faltan columnas obligatorias: {faltantes}"
    try:
        cantidad = float(datos["Cantidad"])
        if cantidad <= 0:
            return "Cantidad debe ser mayor a 0"
    except (TypeError, ValueError):
        return "Cantidad no es un numero valido"
    try:
        costo_unitario = float(datos["Costo unitario"])
        if costo_unitario < 0:
            return "Costo unitario no puede ser negativo"
    except (TypeError, ValueError):
        return "Costo unitario no es un numero valido"
    return None


def _validar_codigos_internos_duplicados(filas: list[dict]) -> dict[int, str]:
    """Si el Excel trae 'Codigo interno' explicito, no puede repetirse
    dentro del mismo archivo (dentro del mismo inventario)."""
    vistos: dict[str, int] = {}
    errores: dict[int, str] = {}
    for f in filas:
        codigo = f["datos"].get("Codigo interno")
        if not codigo:
            continue
        codigo = str(codigo).strip()
        if codigo in vistos:
            errores[f["numero_fila"]] = (
                f"Codigo interno '{codigo}' duplicado en el archivo (fila {vistos[codigo]})"
            )
        else:
            vistos[codigo] = f["numero_fila"]
    return errores


def previsualizar(
    db: Session, inventario_id: int, nombre_archivo: str, contenido: bytes
) -> schemas.CargaPreviewOut:
    """Lee y valida el Excel SIN grabar nada en m02/m03. Devuelve el
    resumen y el detalle de errores por fila para que el usuario confirme."""
    validators.validar_extension_excel(nombre_archivo)
    inventario_service.obtener_inventario(db, inventario_id)  # valida existencia

    try:
        filas = _leer_filas_excel(contenido)
    except ValueError as exc:
        raise ValueError(str(exc)) from exc

    duplicados = _validar_codigos_internos_duplicados(filas)

    carga = CargaInventarioInicial(
        nombre_archivo=nombre_archivo,
        inventario_id=inventario_id,
        estado="PREVISUALIZADA",
        total_filas=len(filas),
    )
    carga = repository.crear_carga(db, carga)

    errores: list[schemas.FilaErrorOut] = []
    validas = 0
    for f in filas:
        numero_fila = f["numero_fila"]
        mensaje_error = duplicados.get(numero_fila) or _validar_fila(f["datos"])
        es_valida = mensaje_error is None
        if es_valida:
            validas += 1
        else:
            errores.append(schemas.FilaErrorOut(numero_fila=numero_fila, mensaje_error=mensaje_error))

        repository.agregar_fila(
            db,
            CargaInventarioInicialFila(
                carga_id=carga.id,
                numero_fila=numero_fila,
                datos_json=json.dumps(f["datos"], ensure_ascii=False),
                valida=es_valida,
                mensaje_error=mensaje_error,
            ),
        )

    def _costo_es_cero(datos: dict) -> bool:
        try:
            return float(datos.get("Costo unitario")) == 0
        except (TypeError, ValueError):
            return False

    filas_costo_cero = sum(
        1
        for f in filas
        if (duplicados.get(f["numero_fila"]) or _validar_fila(f["datos"])) is None
        and _costo_es_cero(f["datos"])
    )

    carga.filas_validas = validas
    carga.filas_con_error = len(errores)
    carga.estado = "CON_ERRORES" if errores else "PREVISUALIZADA"
    carga = repository.guardar_carga(db, carga)

    return schemas.CargaPreviewOut(
        carga_id=carga.id,
        nombre_archivo=carga.nombre_archivo,
        inventario_id=carga.inventario_id,
        estado=carga.estado,
        total_filas=carga.total_filas,
        filas_validas=carga.filas_validas,
        filas_con_error=carga.filas_con_error,
        filas_costo_cero=filas_costo_cero,
        errores=errores,
    )


def obtener_carga(db: Session, carga_id: int) -> CargaInventarioInicial:
    return validators.validar_carga_existe(repository.obtener_carga(db, carga_id))


def listar_cargas(db: Session) -> list[CargaInventarioInicial]:
    return repository.listar_cargas(db)


def _normalizar_nombre_producto(nombre: str) -> str:
    """Colapsa espacios multiples/saltos de linea a uno solo, recorta los
    extremos. Usado tanto para guardar el nombre como para compararlo."""
    return " ".join(str(nombre).split())


def _obtener_o_crear_producto(db: Session, nombre: str) -> Producto:
    """Busca un producto por nombre de forma case-insensitive y tolerando
    espacios extra (fix 2026-08-02): antes comparaba con `==` exacto, asi
    que "BORZINED", "Borzined " y "borzined" (o cualquier variacion de
    mayusculas/espacios del mismo producto en distintas filas/archivos)
    creaban productos DUPLICADOS con stock separado, en vez de acumularse
    en el mismo producto. Esto NO fusiona automaticamente los productos
    duplicados que ya existian en la base de datos antes de este fix --
    eso requiere una limpieza manual aparte, con el usuario decidiendo
    cual es el producto "bueno" en cada caso."""
    nombre_normalizado = _normalizar_nombre_producto(nombre)
    existente = (
        db.query(Producto)
        .filter(func.lower(Producto.nombre) == nombre_normalizado.lower())
        .first()
    )
    if existente is not None:
        return existente
    codigo_auto = f"AUTO-{nombre_normalizado[:20].upper().replace(' ', '-')}"
    codigo_final = codigo_auto
    contador = 1
    while productos_repo.obtener_por_codigo(db, codigo_final) is not None:
        contador += 1
        codigo_final = f"{codigo_auto}-{contador}"
    return productos_repo.crear(
        db, Producto(codigo=codigo_final, nombre=nombre_normalizado, unidad_medida="UND")
    )


def _procesar_fila(db: Session, carga: CargaInventarioInicial, datos: dict, numero_fila: int) -> None:
    producto = _obtener_o_crear_producto(db, str(datos["Producto"]).strip())

    litros = datos.get("Litros Presentacion")
    inventario_service.registrar_ingreso(
        db,
        inv_schemas.IngresoInventarioCrear(
            producto_id=producto.id,
            inventario_id=carga.inventario_id,
            codigo_lote=str(datos["Lote"]).strip(),
            cantidad=float(datos["Cantidad"]),
            costo_unitario=float(datos["Costo unitario"]),
            fecha_elaboracion=datos.get("Fecha elaboracion"),
            fecha_vencimiento=datos.get("Fecha vencimiento"),
            referencia=(
                f"Carga inventario inicial #{carga.id} "
                f"(archivo {carga.nombre_archivo}, fila {numero_fila})"
            ),
            codigo_interno=str(datos["Codigo interno"]).strip() if datos.get("Codigo interno") else None,
            familia=str(datos["Familia"]).strip() if datos.get("Familia") else None,
            presentacion=str(datos["Presentacion"]).strip() if datos.get("Presentacion") else None,
            litros_presentacion=float(litros) if litros not in (None, "") else None,
            marca=None,
        ),
    )


def confirmar(db: Session, carga_id: int) -> schemas.CargaConfirmarOut:
    carga = obtener_carga(db, carga_id)
    validators.validar_carga_no_confirmada(carga)

    fallidas: list[schemas.FilaErrorOut] = []
    procesadas = 0
    for fila in repository.filas_pendientes(db, carga_id):
        datos = json.loads(fila.datos_json)
        try:
            _procesar_fila(db, carga, datos, fila.numero_fila)
            fila.procesada = True
            repository.guardar_fila(db, fila)
            procesadas += 1
        except Exception as exc:
            db.rollback()  # deja la sesion limpia para poder seguir con la fila siguiente
            fallidas.append(schemas.FilaErrorOut(numero_fila=fila.numero_fila, mensaje_error=str(exc)))

    carga.estado = "CONFIRMADA"
    carga.confirmado_en = datetime.now(timezone.utc)
    repository.guardar_carga(db, carga)

    return schemas.CargaConfirmarOut(
        carga_id=carga.id,
        estado=carga.estado,
        filas_procesadas=procesadas,
        filas_fallidas_en_confirmacion=fallidas,
    )


# =======================================================================
# ETAPA 2: fecha de corte + carga historica de Compras y Ventas
# =======================================================================


def configurar_corte_inventario(
    db: Session, inventario_id: int, fecha_corte: datetime
) -> ConfiguracionCorteInventario:
    inventario_service.obtener_inventario(db, inventario_id)  # valida existencia
    corte = repository.obtener_corte_por_inventario(db, inventario_id)
    if corte is None:
        corte = ConfiguracionCorteInventario(inventario_id=inventario_id, fecha_corte=fecha_corte)
    else:
        corte.fecha_corte = fecha_corte
    return repository.guardar_corte(db, corte)


def obtener_corte_inventario(db: Session, inventario_id: int) -> ConfiguracionCorteInventario:
    corte = repository.obtener_corte_por_inventario(db, inventario_id)
    validators.validar_corte_configurado(corte)
    return corte


def _determinar_modo_carga(datos: dict, fecha_corte: datetime | None) -> str:
    """Prioridad: columna explicita tipo_movimiento/modo_carga en el
    Excel. Si no viene, compara la Fecha de la fila contra la fecha de
    corte configurada (solo se compara la parte de fecha, para evitar
    problemas de zona horaria entre un dato naive del Excel y el campo
    con timezone de ConfiguracionCorteInventario)."""
    valor_columna = datos.get("tipo_movimiento")
    if valor_columna not in (None, ""):
        return validators.validar_modo_carga_excel(valor_columna)

    if fecha_corte is None:
        raise ValueError(
            "Falta la columna 'tipo_movimiento'/'modo_carga' y no hay fecha de corte "
            "configurada para este inventario."
        )

    fecha_mov = datos.get("Fecha")
    if fecha_mov in (None, ""):
        raise ValueError("Falta la columna 'Fecha', necesaria para determinar HISTORICO/OPERATIVO")
    try:
        fecha_mov_dt = datetime.fromisoformat(str(fecha_mov))
    except ValueError as exc:
        raise ValueError(f"Fecha invalida: '{fecha_mov}'") from exc

    return "HISTORICO" if fecha_mov_dt.date() <= fecha_corte.date() else "OPERATIVO"


# --- lectura generica de Excel por alias de columnas (no modifica los
#     helpers usados por Inventario Inicial, son funciones nuevas) ---


def _detectar_encabezados_por_alias(
    hoja, alias_columnas: dict[str, str], columnas_obligatorias: list[str]
) -> tuple[int, dict[str, int]]:
    for i, fila in enumerate(hoja.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        indices: dict[str, int] = {}
        for idx, valor in enumerate(fila):
            clave = _clave_encabezado(valor)
            if clave is None:
                continue
            canonica = alias_columnas.get(clave)
            if canonica:
                indices[canonica] = idx
        if all(c in indices for c in columnas_obligatorias):
            return i, indices
    raise ValueError(
        f"No se encontraron los encabezados obligatorios {columnas_obligatorias} "
        f"en las primeras 15 filas de la hoja '{hoja.title}'."
    )


# Unica columna generica de fecha usada por Compras/Ventas historico (el
# resto de valores solo necesitan _normalizar_valor).
_COLUMNAS_FECHA_GENERICO = {"Fecha"}


def _leer_filas_excel_por_alias(
    contenido: bytes, alias_columnas: dict[str, str], columnas_obligatorias: list[str]
) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
    filas_resultado = []
    errores_por_hoja: dict[str, str] = {}
    for hoja in wb.worksheets:
        try:
            fila_encabezado, indices = _detectar_encabezados_por_alias(
                hoja, alias_columnas, columnas_obligatorias
            )
        except ValueError as exc:
            errores_por_hoja[hoja.title] = str(exc)
            continue
        _expandir_celdas_combinadas(hoja)  # soporte para Compras/Ventas con celdas combinadas
        for num_fila, fila in enumerate(
            hoja.iter_rows(min_row=fila_encabezado + 1, values_only=True), start=fila_encabezado + 1
        ):
            valores = {}
            for col, idx in indices.items():
                if idx >= len(fila):
                    continue
                valor = fila[idx]
                valores[col] = (
                    _normalizar_fecha(valor) if col in _COLUMNAS_FECHA_GENERICO else _normalizar_valor(valor)
                )
            if not any(v not in (None, "") for v in valores.values()):
                continue
            valores["_hoja"] = hoja.title
            filas_resultado.append({"numero_fila": num_fila, "datos": valores})

    if not filas_resultado and errores_por_hoja:
        # Mismo bug/fix que _leer_filas_excel: si ninguna hoja del libro
        # tiene los encabezados obligatorios, se avisa explicitamente en
        # vez de devolver un preview "exitoso" con 0 filas.
        detalle = "; ".join(f"'{h}': {m}" for h, m in errores_por_hoja.items())
        raise ValueError(
            f"No se pudo leer ninguna fila del Excel: ninguna de sus hojas tiene "
            f"los encabezados obligatorios {columnas_obligatorias}. Detalle por hoja: {detalle}"
        )
    return filas_resultado


def _obtener_o_crear_proveedor(db: Session, ruc: str, razon_social: str) -> Proveedor:
    existente = proveedores_repo.obtener_por_ruc(db, ruc)
    if existente is not None:
        return existente
    return proveedores_repo.crear(db, Proveedor(ruc=ruc, razon_social=razon_social))


def _obtener_o_crear_cliente(db: Session, ruc: str, razon_social: str) -> Cliente:
    existente = clientes_repo.obtener_por_ruc(db, ruc)
    if existente is not None:
        return existente
    return clientes_repo.crear(db, Cliente(ruc=ruc, razon_social=razon_social))


# -----------------------------------------------------------------
# COMPRAS HISTORICO
# -----------------------------------------------------------------

COLUMNAS_OBLIGATORIAS_COMPRAS = [
    "RUC Proveedor",
    "Razon Social Proveedor",
    "Producto",
    "Cantidad",
    "Costo unitario",
    "Fecha",
]

ALIAS_COLUMNAS_COMPRAS = {
    "ruc proveedor": "RUC Proveedor",
    "ruc": "RUC Proveedor",
    "razon social proveedor": "Razon Social Proveedor",
    "razon social": "Razon Social Proveedor",
    "proveedor": "Razon Social Proveedor",  # header real del cliente: "Proveedor"
    "producto": "Producto",
    "descripcion": "Producto",  # header real del cliente: "Descripcion"
    "concepto": "Producto",  # header real del cliente: "CONCEPTO" (nombre normalizado sin presentacion/unidad mezclada)
    "cantidad": "Cantidad",
    "costo unitario": "Costo unitario",
    "precio de compra": "Costo unitario",  # header real del cliente
    "precio compra": "Costo unitario",
    "fecha": "Fecha",
    "fecha de emision": "Fecha",  # header real del cliente: "Fecha de Emision Factura"
    "fecha de emision factura": "Fecha",
    "fecha emision": "Fecha",
    "moneda": "Moneda",
    "nro documento": "Nro documento",
    "numero documento": "Nro documento",
    "factura": "Nro documento",  # header real del cliente: "Factura"
    "lote": "Lote",
    "tipo movimiento": "tipo_movimiento",
    "tipo_movimiento": "tipo_movimiento",
    "modo carga": "tipo_movimiento",
    "modo_carga": "tipo_movimiento",
    # --- Fix auditoria 2026-08-03: el importador historico ignoraba estas
    # columnas del Excel del cliente (COMPRAS_ECO_NEOAGROX_2026), aun cuando
    # veian llenas fila por fila (ver OC #81, fila 84). Mismos alias que ya
    # usa m04_compras/importacion_service.py para el importador "normal".
    "presentacion": "Presentacion",
    "unidad de medida": "Unidad Medida",
    "unida de medida": "Unidad Medida",  # tal como viene escrito en el Excel del cliente
    "cantidad por unidad": "Cantidad Por Unidad",
}


def _texto_opcional(valor) -> str | None:
    if valor is None:
        return None
    texto = str(valor).strip()
    return texto or None


def _es_numero_opcional(valor) -> bool:
    if valor in (None, ""):
        return False
    try:
        float(valor)
        return True
    except (TypeError, ValueError):
        return False


def _validar_fila_compra(datos: dict) -> str | None:
    faltantes = [c for c in COLUMNAS_OBLIGATORIAS_COMPRAS if not datos.get(c)]
    if faltantes:
        return f"Faltan columnas obligatorias: {faltantes}"
    try:
        cantidad = float(datos["Cantidad"])
        if cantidad <= 0:
            return "Cantidad debe ser mayor a 0"
    except (TypeError, ValueError):
        return "Cantidad no es un numero valido"
    try:
        costo = float(datos["Costo unitario"])
        if costo < 0:
            return "Costo unitario no puede ser negativo"
    except (TypeError, ValueError):
        return "Costo unitario no es un numero valido"
    return None


def previsualizar_compras(
    db: Session, inventario_id: int, nombre_archivo: str, contenido: bytes
) -> schemas.CargaComprasPreviewOut:
    validators.validar_extension_excel(nombre_archivo)
    inventario_service.obtener_inventario(db, inventario_id)  # valida existencia

    corte = repository.obtener_corte_por_inventario(db, inventario_id)
    fecha_corte = corte.fecha_corte if corte is not None else None

    try:
        filas = _leer_filas_excel_por_alias(
            contenido, ALIAS_COLUMNAS_COMPRAS, COLUMNAS_OBLIGATORIAS_COMPRAS
        )
    except ValueError as exc:
        raise ValueError(str(exc)) from exc

    carga = CargaComprasHistorico(
        nombre_archivo=nombre_archivo, inventario_id=inventario_id, total_filas=len(filas)
    )
    carga = repository.crear_carga_compras(db, carga)

    errores: list[schemas.FilaErrorOut] = []
    validas = 0
    historico = 0
    operativo = 0
    for f in filas:
        numero_fila = f["numero_fila"]
        mensaje_error = _validar_fila_compra(f["datos"])
        modo_carga = None
        if mensaje_error is None:
            try:
                modo_carga = _determinar_modo_carga(f["datos"], fecha_corte)
            except ValueError as exc:
                mensaje_error = str(exc)

        es_valida = mensaje_error is None
        if es_valida:
            validas += 1
            if modo_carga == "HISTORICO":
                historico += 1
            else:
                operativo += 1
        else:
            errores.append(schemas.FilaErrorOut(numero_fila=numero_fila, mensaje_error=mensaje_error))

        repository.agregar_fila_compras(
            db,
            CargaComprasHistoricoFila(
                carga_id=carga.id,
                numero_fila=numero_fila,
                datos_json=json.dumps(f["datos"], ensure_ascii=False),
                modo_carga=modo_carga or "OPERATIVO",  # placeholder si la fila es invalida
                valida=es_valida,
                mensaje_error=mensaje_error,
            ),
        )

    carga.filas_validas = validas
    carga.filas_con_error = len(errores)
    carga.estado = "CON_ERRORES" if errores else "PREVISUALIZADA"
    carga = repository.guardar_carga_compras(db, carga)

    return schemas.CargaComprasPreviewOut(
        carga_id=carga.id,
        nombre_archivo=carga.nombre_archivo,
        inventario_id=carga.inventario_id,
        estado=carga.estado,
        total_filas=carga.total_filas,
        filas_validas=carga.filas_validas,
        filas_con_error=carga.filas_con_error,
        filas_historico=historico,
        filas_operativo=operativo,
        errores=errores,
    )


def obtener_carga_compras(db: Session, carga_id: int) -> CargaComprasHistorico:
    return validators.validar_carga_existe(repository.obtener_carga_compras(db, carga_id))


def listar_cargas_compras(db: Session) -> list[CargaComprasHistorico]:
    return repository.listar_cargas_compras(db)


def _procesar_fila_compra(
    db: Session, carga: CargaComprasHistorico, datos: dict, numero_fila: int, modo_carga: str
) -> int:
    proveedor = _obtener_o_crear_proveedor(
        db, str(datos["RUC Proveedor"]).strip(), str(datos["Razon Social Proveedor"]).strip()
    )
    producto = _obtener_o_crear_producto(db, str(datos["Producto"]).strip())
    fecha_movimiento = datetime.fromisoformat(str(datos["Fecha"]))

    orden = compras_service.crear_orden(
        db,
        compras_schemas.OrdenCompraCrear(
            proveedor_id=proveedor.id,
            inventario_destino_id=carga.inventario_id,
            moneda=_normalizar_moneda(datos.get("Moneda"), "USD"),
            observaciones=(
                f"Carga compras historico #{carga.id} (archivo {carga.nombre_archivo}, "
                f"fila {numero_fila}, doc {datos.get('Nro documento') or 's/n'})"
            ),
            items=[
                compras_schemas.OrdenCompraItemCrear(
                    producto_id=producto.id,
                    cantidad=float(datos["Cantidad"]),
                    costo_unitario=float(datos["Costo unitario"]),
                    presentacion=_texto_opcional(datos.get("Presentacion")),
                    unidad_medida=_texto_opcional(datos.get("Unidad Medida")),
                    cantidad_por_unidad=(
                        float(datos["Cantidad Por Unidad"])
                        if _es_numero_opcional(datos.get("Cantidad Por Unidad"))
                        else None
                    ),
                )
            ],
            # Paso 2 (Decision 1): creado_en de la orden respeta la fecha
            # del Excel, sin importar el modo_carga (documental u operativo).
            fecha_movimiento=fecha_movimiento,
        ),
    )
    orden = compras_service.aprobar_orden(db, orden.id, fecha_movimiento=fecha_movimiento)

    if modo_carga == "HISTORICO":
        # Solo documental: se fija el estado final directo por repository,
        # SIN pasar por compras_service.recibir_orden() (esa si mueve Kardex/stock).
        orden.estado = "RECIBIDA"
        orden.recibido_en = fecha_movimiento
        orden = compras_repo.guardar(db, orden)
    else:
        # OPERATIVO: movimiento real, reutiliza el servicio existente de m04
        # tal cual, que a su vez llama a m03.registrar_ingreso. Paso 2: se
        # propaga fecha_movimiento para que recibido_en y el Kardex/Lote
        # que nace en m03 queden con la fecha real del Excel (y el FEFO
        # evalue vencimiento contra ella), no con la fecha de carga.
        orden = compras_service.recibir_orden(db, orden.id, fecha_movimiento=fecha_movimiento)

    return orden.id


def confirmar_compras(db: Session, carga_id: int) -> schemas.CargaComprasConfirmarOut:
    carga = obtener_carga_compras(db, carga_id)
    validators.validar_carga_no_confirmada(carga)

    fallidas: list[schemas.FilaErrorOut] = []
    procesadas = 0
    historico_creadas = 0
    operativas_creadas = 0
    for fila in repository.filas_pendientes_compras(db, carga_id):
        datos = json.loads(fila.datos_json)
        try:
            fila.orden_compra_id = _procesar_fila_compra(db, carga, datos, fila.numero_fila, fila.modo_carga)
            fila.procesada = True
            repository.guardar_fila_compras(db, fila)
            procesadas += 1
            if fila.modo_carga == "HISTORICO":
                historico_creadas += 1
            else:
                operativas_creadas += 1
        except Exception as exc:
            db.rollback()
            fallidas.append(schemas.FilaErrorOut(numero_fila=fila.numero_fila, mensaje_error=str(exc)))

    carga.estado = "CONFIRMADA"
    carga.confirmado_en = datetime.now(timezone.utc)
    repository.guardar_carga_compras(db, carga)

    return schemas.CargaComprasConfirmarOut(
        carga_id=carga.id,
        estado=carga.estado,
        filas_procesadas=procesadas,
        filas_historico_creadas=historico_creadas,
        filas_operativas_creadas=operativas_creadas,
        filas_fallidas_en_confirmacion=fallidas,
    )


# -----------------------------------------------------------------
# VENTAS HISTORICO
# -----------------------------------------------------------------

COLUMNAS_OBLIGATORIAS_VENTAS = [
    "RUC Cliente",
    "Razon Social Cliente",
    "Producto",
    "Cantidad",
    "Precio unitario",
    "Fecha",
]

ALIAS_COLUMNAS_VENTAS = {
    "ruc cliente": "RUC Cliente",
    "ruc": "RUC Cliente",
    "razon social cliente": "Razon Social Cliente",
    "razon social": "Razon Social Cliente",
    "cliente": "Razon Social Cliente",  # header real del cliente: "Cliente"
    "producto": "Producto",
    "descripcion": "Producto",  # header real del cliente: "Descripcion"
    "concepto": "Producto",  # header real del cliente: "CONCEPTO" (nombre normalizado sin presentacion/unidad mezclada)
    "cantidad": "Cantidad",
    "precio unitario": "Precio unitario",
    "precio venta": "Precio unitario",  # header real del cliente: "Precio Venta"
    "precio de venta": "Precio unitario",
    "fecha": "Fecha",
    "fecha de emision": "Fecha",  # header real del cliente: "Fecha de Emision"
    "fecha emision": "Fecha",
    "moneda": "Moneda",
    "nro documento": "Nro documento",
    "numero documento": "Nro documento",
    "factura": "Nro documento",  # header real del cliente: "Factura"
    "tipo movimiento": "tipo_movimiento",
    "tipo_movimiento": "tipo_movimiento",
    "modo carga": "tipo_movimiento",
    "modo_carga": "tipo_movimiento",
    # --- Fix auditoria 2026-08-03: mismo bug que en COMPRAS_HISTORICO, el
    # importador ignoraba esta columna del Excel del cliente aunque venia
    # llena fila por fila (ver VENTAS_2026 - INVENTARIO, columna "Unida de
    # Medida"). Ventas no trae "Presentacion" como columna propia, solo
    # unidad de medida.
    "unidad de medida": "Unidad de Medida",
    "unida de medida": "Unidad de Medida",  # tal como viene escrito en el Excel del cliente
}


def _validar_fila_venta(datos: dict) -> str | None:
    faltantes = [c for c in COLUMNAS_OBLIGATORIAS_VENTAS if not datos.get(c)]
    if faltantes:
        return f"Faltan columnas obligatorias: {faltantes}"
    try:
        cantidad = float(datos["Cantidad"])
        if cantidad <= 0:
            return "Cantidad debe ser mayor a 0"
    except (TypeError, ValueError):
        return "Cantidad no es un numero valido"
    try:
        precio = float(datos["Precio unitario"])
        if precio < 0:
            return "Precio unitario no puede ser negativo"
    except (TypeError, ValueError):
        return "Precio unitario no es un numero valido"
    return None


def previsualizar_ventas(
    db: Session, inventario_id: int, nombre_archivo: str, contenido: bytes
) -> schemas.CargaVentasPreviewOut:
    validators.validar_extension_excel(nombre_archivo)
    inventario_service.obtener_inventario(db, inventario_id)  # valida existencia

    corte = repository.obtener_corte_por_inventario(db, inventario_id)
    fecha_corte = corte.fecha_corte if corte is not None else None

    try:
        filas = _leer_filas_excel_por_alias(
            contenido, ALIAS_COLUMNAS_VENTAS, COLUMNAS_OBLIGATORIAS_VENTAS
        )
    except ValueError as exc:
        raise ValueError(str(exc)) from exc

    carga = CargaVentasHistorico(
        nombre_archivo=nombre_archivo, inventario_id=inventario_id, total_filas=len(filas)
    )
    carga = repository.crear_carga_ventas(db, carga)

    errores: list[schemas.FilaErrorOut] = []
    validas = 0
    historico = 0
    operativo = 0
    for f in filas:
        numero_fila = f["numero_fila"]
        mensaje_error = _validar_fila_venta(f["datos"])
        modo_carga = None
        if mensaje_error is None:
            try:
                modo_carga = _determinar_modo_carga(f["datos"], fecha_corte)
            except ValueError as exc:
                mensaje_error = str(exc)

        es_valida = mensaje_error is None
        if es_valida:
            validas += 1
            if modo_carga == "HISTORICO":
                historico += 1
            else:
                operativo += 1
        else:
            errores.append(schemas.FilaErrorOut(numero_fila=numero_fila, mensaje_error=mensaje_error))

        repository.agregar_fila_ventas(
            db,
            CargaVentasHistoricoFila(
                carga_id=carga.id,
                numero_fila=numero_fila,
                datos_json=json.dumps(f["datos"], ensure_ascii=False),
                modo_carga=modo_carga or "OPERATIVO",
                valida=es_valida,
                mensaje_error=mensaje_error,
            ),
        )

    carga.filas_validas = validas
    carga.filas_con_error = len(errores)
    carga.estado = "CON_ERRORES" if errores else "PREVISUALIZADA"
    carga = repository.guardar_carga_ventas(db, carga)

    return schemas.CargaVentasPreviewOut(
        carga_id=carga.id,
        nombre_archivo=carga.nombre_archivo,
        inventario_id=carga.inventario_id,
        estado=carga.estado,
        total_filas=carga.total_filas,
        filas_validas=carga.filas_validas,
        filas_con_error=carga.filas_con_error,
        filas_historico=historico,
        filas_operativo=operativo,
        errores=errores,
    )


def obtener_carga_ventas(db: Session, carga_id: int) -> CargaVentasHistorico:
    return validators.validar_carga_existe(repository.obtener_carga_ventas(db, carga_id))


def listar_cargas_ventas(db: Session) -> list[CargaVentasHistorico]:
    return repository.listar_cargas_ventas(db)


def _procesar_fila_venta(
    db: Session, carga: CargaVentasHistorico, datos: dict, numero_fila: int, modo_carga: str
) -> int:
    cliente = _obtener_o_crear_cliente(
        db, str(datos["RUC Cliente"]).strip(), str(datos["Razon Social Cliente"]).strip()
    )
    producto = _obtener_o_crear_producto(db, str(datos["Producto"]).strip())
    fecha_movimiento = datetime.fromisoformat(str(datos["Fecha"]))

    orden = ventas_service.crear_orden(
        db,
        ventas_schemas.OrdenVentaCrear(
            cliente_id=cliente.id,
            inventario_salida_id=carga.inventario_id,
            moneda=_normalizar_moneda(datos.get("Moneda"), "PEN"),
            observaciones=(
                f"Carga ventas historico #{carga.id} (archivo {carga.nombre_archivo}, "
                f"fila {numero_fila}, doc {datos.get('Nro documento') or 's/n'})"
            ),
            items=[
                ventas_schemas.OrdenVentaItemCrear(
                    producto_id=producto.id,
                    cantidad=float(datos["Cantidad"]),
                    precio_unitario_venta=float(datos["Precio unitario"]),
                    unidad_medida=_texto_opcional(datos.get("Unidad de Medida")),
                )
            ],
            # Paso 2 (Decision 1): creado_en de la orden respeta la fecha
            # del Excel, sin importar el modo_carga (documental u operativo).
            fecha_movimiento=fecha_movimiento,
        ),
    )
    orden = ventas_service.confirmar_orden(db, orden.id, fecha_movimiento=fecha_movimiento)

    if modo_carga == "HISTORICO":
        # Solo documental: se fija el estado final directo por repository,
        # SIN pasar por ventas_service.despachar_orden() (esa si descuenta Kardex/stock).
        orden.estado = "DESPACHADA"
        orden.despachado_en = fecha_movimiento
        orden = ventas_repo.guardar(db, orden)
    else:
        # OPERATIVO: movimiento real (descuenta stock por FEFO via m03),
        # reutiliza el servicio existente de m10 tal cual. Paso 2: se
        # propaga fecha_movimiento para que despachado_en y el Kardex que
        # nace en m03 queden con la fecha real del Excel (y el FEFO evalue
        # vencimiento contra ella), no con la fecha de carga.
        orden = ventas_service.despachar_orden(db, orden.id, fecha_movimiento=fecha_movimiento)

    return orden.id


def confirmar_ventas(db: Session, carga_id: int) -> schemas.CargaVentasConfirmarOut:
    carga = obtener_carga_ventas(db, carga_id)
    validators.validar_carga_no_confirmada(carga)

    fallidas: list[schemas.FilaErrorOut] = []
    procesadas = 0
    historico_creadas = 0
    operativas_creadas = 0
    for fila in repository.filas_pendientes_ventas(db, carga_id):
        datos = json.loads(fila.datos_json)
        try:
            fila.orden_venta_id = _procesar_fila_venta(db, carga, datos, fila.numero_fila, fila.modo_carga)
            fila.procesada = True
            repository.guardar_fila_ventas(db, fila)
            procesadas += 1
            if fila.modo_carga == "HISTORICO":
                historico_creadas += 1
            else:
                operativas_creadas += 1
        except Exception as exc:
            db.rollback()
            fallidas.append(schemas.FilaErrorOut(numero_fila=fila.numero_fila, mensaje_error=str(exc)))

    carga.estado = "CONFIRMADA"
    carga.confirmado_en = datetime.now(timezone.utc)
    repository.guardar_carga_ventas(db, carga)

    return schemas.CargaVentasConfirmarOut(
        carga_id=carga.id,
        estado=carga.estado,
        filas_procesadas=procesadas,
        filas_historico_creadas=historico_creadas,
        filas_operativas_creadas=operativas_creadas,
        filas_fallidas_en_confirmacion=fallidas,
    )


# =======================================================================
# ETAPA 3: REEMPLAZAR IMPORTACION CONFIRMADA
# (Inventario Inicial / Compras historico / Ventas historico)
#
# Regla de oro para no perder datos: SIEMPRE se previsualiza el archivo
# NUEVO antes de tocar un solo registro de la carga vieja. Si el archivo
# nuevo tiene errores, se aborta sin borrar nada (la previsualizacion
# fallida queda solo como un intento suelto, auditable, que no toca la
# carga original).
#
# La carga vieja (fila de cargas_inventario_inicial / cargas_compras_
# historico / cargas_ventas_historico) NUNCA se borra: solo se marca
# estado_vigencia="REEMPLAZADA" y queda enlazada a la carga nueva para
# auditoria. Lo que si se elimina, y solo dentro de la transaccion
# atomica y solo si el archivo nuevo es 100% valido, son los registros
# derivados que esa carga genero (Kardex, Lotes, Ordenes de Compra/Venta):
# son artefactos operativos, no la carga en si, y quedan totalmente
# documentados en la bitacora (cantidades, motivo, usuario, fecha).
#
# Todo el reemplazo corre en UNA sola transaccion SQLAlchemy real: si
# cualquier paso falla, ROLLBACK COMPLETO y no queda nada a medias.
# =======================================================================


@contextmanager
def _transaccion_atomica(db: Session):
    """Unica transaccion SQLAlchemy real para todo el reemplazo.

    El resto del modulo (y de m03/m04/m10, que reutilizamos tal cual)
    sigue el patron "eager commit" ya existente: cada funcion de
    repository hace su propio db.add()+db.commit()+db.refresh(). Eso es
    perfecto para su uso normal, pero si se llama tal cual dentro de un
    reemplazo, cada commit() intermedio dejaria el reemplazo a medio
    camino si algo posterior fallara.

    Para lograr atomicidad real SIN tocar ni un solo modulo mas (m02,
    m03, m04, m10 quedan exactamente igual), este context manager
    reemplaza temporalmente el metodo commit() de la sesion por un
    simple flush(): la fila se ve reflejada de inmediato para las
    siguientes consultas del propio proceso, pero fisicamente NO se
    confirma en disco hasta el commit() real del final de este bloque.

    Si ocurre cualquier excepcion, se hace rollback() real -- que
    deshace absolutamente todo lo que paso desde que se entro al
    bloque, porque nada se habia confirmado todavia -- y se relanza la
    excepcion tal cual para que el llamador decida que responder.
    """
    commit_real = db.commit
    db.commit = db.flush
    try:
        yield db
        db.commit = commit_real
        commit_real()
    except Exception:
        db.commit = commit_real
        db.rollback()
        raise


def _bitacora_error(
    db: Session,
    tipo_carga: str,
    carga_anterior_id: int,
    usuario,
    motivo,
    observaciones,
    ip,
    resultado: str,
    detalle: str,
    tiempo_ejecucion_ms: int,
) -> None:
    """Registra en bitacora un intento de reemplazo que NO se completo
    (bloqueado o con error/rollback). Se hace en su PROPIA transaccion
    (fuera de cualquier bloque atomico ya deshecho), para que el intento
    fallido tambien quede auditado: 'resultado' es un campo obligatorio
    de la bitacora, exitoso o no."""
    try:
        repository.crear_bitacora(
            db,
            BitacoraReemplazo(
                tipo_carga=tipo_carga,
                carga_anterior_id=carga_anterior_id,
                carga_nueva_id=None,
                usuario_id=getattr(usuario, "id", None),
                usuario_username=getattr(usuario, "username", "desconocido"),
                ip_origen=ip,
                motivo=motivo or "(sin motivo: bloqueado antes de validar motivo)",
                observaciones=observaciones,
                cantidad_lotes_eliminados=0,
                cantidad_kardex_eliminados=0,
                cantidad_ordenes_eliminadas=0,
                cantidad_registros_nuevos=0,
                tiempo_ejecucion_ms=tiempo_ejecucion_ms,
                resultado=resultado,
                detalle=detalle,
            ),
        )
    except Exception:
        # La bitacora es auditoria: nunca debe tapar el error real del
        # reemplazo ni dejar la sesion en mal estado si algo raro pasara.
        db.rollback()


def listar_bitacora_reemplazos(db: Session, tipo_carga: str | None = None) -> list[BitacoraReemplazo]:
    return repository.listar_bitacora(db, tipo_carga)


# -----------------------------------------------------------------
# Helpers de bloqueo (solo lectura), compartidos por los 3 tipos
# -----------------------------------------------------------------


def _kardex_por_referencia_like(db: Session, patron: str) -> list[MovimientoKardex]:
    return db.query(MovimientoKardex).filter(MovimientoKardex.referencia.like(patron)).all()


def _kardex_por_referencia_in(db: Session, referencias: list[str]) -> list[MovimientoKardex]:
    if not referencias:
        return []
    return db.query(MovimientoKardex).filter(MovimientoKardex.referencia.in_(referencias)).all()


def _kardex_posteriores(
    db: Session,
    lote_ids,
    referencias_propias: list[str] | None = None,
    patron_propio: str | None = None,
) -> list[MovimientoKardex]:
    """Movimientos de Kardex sobre esos lotes que NO son los que genero
    la propia carga (es decir: consumo/ajuste/otro ingreso posterior por
    otro proceso). Si existen, bloquean el reemplazo."""
    if not lote_ids:
        return []
    query = db.query(MovimientoKardex).filter(MovimientoKardex.lote_id.in_(lote_ids))
    if patron_propio is not None:
        query = query.filter(~MovimientoKardex.referencia.like(patron_propio))
    elif referencias_propias:
        query = query.filter(~MovimientoKardex.referencia.in_(referencias_propias))
    return query.all()


# -----------------------------------------------------------------
# INVENTARIO INICIAL
# -----------------------------------------------------------------


def verificar_reemplazo_inventario(db: Session, carga_id: int) -> schemas.ValidacionReemplazoOut:
    carga = obtener_carga(db, carga_id)
    bloqueos: list[schemas.BloqueoReemplazoOut] = []

    if carga.estado != "CONFIRMADA":
        bloqueos.append(schemas.BloqueoReemplazoOut(
            tipo="ESTADO",
            detalle=f"La carga esta en estado '{carga.estado}', debe estar CONFIRMADA para poder reemplazarse.",
        ))
    if carga.estado_vigencia == "REEMPLAZADA":
        bloqueos.append(schemas.BloqueoReemplazoOut(
            tipo="YA_REEMPLAZADA",
            detalle=f"Esta carga ya fue reemplazada por la carga #{carga.carga_reemplazo_id}.",
        ))

    patron = f"Carga inventario inicial #{carga.id} %"
    movimientos_propios = _kardex_por_referencia_like(db, patron)
    lote_ids = {m.lote_id for m in movimientos_propios}
    posteriores = _kardex_posteriores(db, lote_ids, patron_propio=patron)
    if posteriores:
        referencias = sorted({m.referencia or "(sin referencia)" for m in posteriores})
        bloqueos.append(schemas.BloqueoReemplazoOut(
            tipo="MOVIMIENTOS_POSTERIORES",
            detalle=(
                f"{len(posteriores)} movimiento(s) de kardex posterior(es) (ingresos/salidas/ajustes -- "
                "compras, ventas o costeo ya aplicados sobre estos lotes) generados por: "
                + "; ".join(referencias)
            ),
        ))

    if lote_ids and not posteriores:
        lotes = db.query(Lote).filter(Lote.id.in_(lote_ids)).all()
        consumidos = [l for l in lotes if float(l.cantidad_actual) != float(l.cantidad_inicial)]
        if consumidos:
            bloqueos.append(schemas.BloqueoReemplazoOut(
                tipo="STOCK_CONSUMIDO",
                detalle=f"{len(consumidos)} lote(s) de esta carga ya tienen saldo distinto al inicial (hubo consumo).",
            ))

    return schemas.ValidacionReemplazoOut(
        carga_id=carga.id,
        tipo_carga="INVENTARIO_INICIAL",
        estado_carga=carga.estado,
        estado_vigencia=carga.estado_vigencia,
        puede_reemplazar=len(bloqueos) == 0,
        bloqueos=bloqueos,
    )


def reemplazar_inventario(
    db: Session,
    carga_id: int,
    nombre_archivo: str,
    contenido: bytes,
    motivo: str,
    observaciones: str | None,
    usuario,
    ip: str | None = None,
) -> schemas.ReemplazoOut:
    inicio = time.monotonic()
    validators.validar_permiso_reemplazo(usuario)
    motivo = validators.validar_motivo_obligatorio(motivo)

    carga = obtener_carga(db, carga_id)
    validators.validar_carga_confirmada(carga)
    validators.validar_carga_no_reemplazada(carga)

    validacion = verificar_reemplazo_inventario(db, carga_id)
    if not validacion.puede_reemplazar:
        tiempo_ms = int((time.monotonic() - inicio) * 1000)
        _bitacora_error(
            db, "INVENTARIO_INICIAL", carga.id, usuario, motivo, observaciones, ip,
            "BLOQUEADO", "; ".join(f"[{b.tipo}] {b.detalle}" for b in validacion.bloqueos), tiempo_ms,
        )
    validators.validar_sin_bloqueos(validacion.bloqueos)

    # 1) Previsualizar el archivo NUEVO primero, FUERA de la transaccion
    #    atomica: si tiene errores, se aborta aqui mismo sin haber tocado
    #    la carga vieja todavia.
    try:
        preview = previsualizar(db, carga.inventario_id, nombre_archivo, contenido)
    except ValueError as exc:
        tiempo_ms = int((time.monotonic() - inicio) * 1000)
        _bitacora_error(db, "INVENTARIO_INICIAL", carga.id, usuario, motivo, observaciones, ip, "ERROR", str(exc), tiempo_ms)
        raise

    if preview.filas_con_error:
        detalle = (
            f"El archivo de reemplazo tiene {preview.filas_con_error} fila(s) con error "
            f"(carga de previsualizacion #{preview.carga_id}, sin confirmar). "
            "Corrija el Excel y vuelva a intentar el reemplazo; la carga original no fue modificada."
        )
        tiempo_ms = int((time.monotonic() - inicio) * 1000)
        _bitacora_error(db, "INVENTARIO_INICIAL", carga.id, usuario, motivo, observaciones, ip, "ERROR", detalle, tiempo_ms)
        raise ValueError(detalle)

    # 2) A partir de aqui, TODO en una sola transaccion atomica real.
    try:
        with _transaccion_atomica(db):
            patron = f"Carga inventario inicial #{carga.id} %"
            movimientos = _kardex_por_referencia_like(db, patron)
            lote_ids = list({m.lote_id for m in movimientos})
            kardex_eliminados = len(movimientos)
            for m in movimientos:
                db.delete(m)
            db.flush()

            lotes_eliminados = 0
            if lote_ids:
                lotes = db.query(Lote).filter(Lote.id.in_(lote_ids)).all()
                lotes_eliminados = len(lotes)
                for l in lotes:
                    db.delete(l)
            db.flush()

            resultado_confirmacion = confirmar(db, preview.carga_id)
            if resultado_confirmacion.filas_fallidas_en_confirmacion:
                # La previsualizacion ya habia validado el archivo entero
                # sin errores; si igual fallo alguna fila al confirmar, no
                # dejamos un reemplazo a medias: se aborta TODO el bloque.
                raise ValueError(
                    "La carga de reemplazo no se pudo confirmar por completo "
                    f"({len(resultado_confirmacion.filas_fallidas_en_confirmacion)} fila(s) fallaron "
                    "pese a haber pasado la previsualizacion). Se revirtio todo el reemplazo, "
                    "la carga original sigue vigente sin cambios."
                )

            nueva_carga = obtener_carga(db, preview.carga_id)
            repository.marcar_carga_reemplazada(db, carga, nueva_carga, getattr(usuario, "id", None), motivo, observaciones)

            tiempo_ms = int((time.monotonic() - inicio) * 1000)
            bitacora = BitacoraReemplazo(
                tipo_carga="INVENTARIO_INICIAL",
                carga_anterior_id=carga.id,
                carga_nueva_id=nueva_carga.id,
                usuario_id=getattr(usuario, "id", None),
                usuario_username=getattr(usuario, "username", "desconocido"),
                ip_origen=ip,
                motivo=motivo,
                observaciones=observaciones,
                cantidad_lotes_eliminados=lotes_eliminados,
                cantidad_kardex_eliminados=kardex_eliminados,
                cantidad_ordenes_eliminadas=0,
                cantidad_registros_nuevos=resultado_confirmacion.filas_procesadas,
                tiempo_ejecucion_ms=tiempo_ms,
                resultado="EXITOSO",
                detalle=None,
            )
            db.add(bitacora)
            db.flush()
            db.refresh(bitacora)
    except Exception as exc:
        tiempo_ms = int((time.monotonic() - inicio) * 1000)
        _bitacora_error(db, "INVENTARIO_INICIAL", carga.id, usuario, motivo, observaciones, ip, "ERROR", str(exc), tiempo_ms)
        raise

    db.refresh(carga)
    return schemas.ReemplazoOut(
        carga_anterior_id=carga.id,
        carga_anterior_estado_vigencia=carga.estado_vigencia,
        carga_nueva_id=nueva_carga.id,
        carga_nueva_estado_vigencia=nueva_carga.estado_vigencia,
        kardex_eliminados=kardex_eliminados,
        lotes_eliminados=lotes_eliminados,
        ordenes_eliminadas=0,
        filas_procesadas=resultado_confirmacion.filas_procesadas,
        filas_fallidas_en_confirmacion=resultado_confirmacion.filas_fallidas_en_confirmacion,
        motivo=motivo,
        tiempo_ejecucion_ms=tiempo_ms,
        bitacora_id=bitacora.id,
    )


# -----------------------------------------------------------------
# COMPRAS HISTORICO
# -----------------------------------------------------------------


def verificar_reemplazo_compras(db: Session, carga_id: int) -> schemas.ValidacionReemplazoOut:
    carga = obtener_carga_compras(db, carga_id)
    bloqueos: list[schemas.BloqueoReemplazoOut] = []

    if carga.estado != "CONFIRMADA":
        bloqueos.append(schemas.BloqueoReemplazoOut(
            tipo="ESTADO",
            detalle=f"La carga esta en estado '{carga.estado}', debe estar CONFIRMADA para poder reemplazarse.",
        ))
    if carga.estado_vigencia == "REEMPLAZADA":
        bloqueos.append(schemas.BloqueoReemplazoOut(
            tipo="YA_REEMPLAZADA",
            detalle=f"Esta carga ya fue reemplazada por la carga #{carga.carga_reemplazo_id}.",
        ))

    filas = repository.filas_procesadas_compras(db, carga_id)
    orden_ids = sorted({f.orden_compra_id for f in filas if f.orden_compra_id is not None})

    if orden_ids:
        costeos = (
            db.query(CostoAdicional)
            .filter(CostoAdicional.tipo_documento == "COMPRA", CostoAdicional.documento_id.in_(orden_ids))
            .all()
        )
        if costeos:
            ordenes_afectadas = sorted({c.documento_id for c in costeos})
            bloqueos.append(schemas.BloqueoReemplazoOut(
                tipo="COSTEO",
                detalle=(
                    f"{len(costeos)} costo(s) adicional(es) (flete/seguro/aduana/almacenaje/...) "
                    f"registrados sobre la(s) orden(es) de compra {ordenes_afectadas}."
                ),
            ))

        operaciones = (
            db.query(OperacionLogistica)
            .filter(OperacionLogistica.orden_compra_id.in_(orden_ids))
            .all()
        )
        if operaciones:
            ordenes_afectadas = sorted({o.orden_compra_id for o in operaciones})
            bloqueos.append(schemas.BloqueoReemplazoOut(
                tipo="OPERACION_LOGISTICA",
                detalle=(
                    f"{len(operaciones)} operacion(es) logistica(s) (recepcion/inspeccion/ubicacion/"
                    f"picking/despacho) registradas sobre la(s) orden(es) de compra {ordenes_afectadas}."
                ),
            ))

    referencias_propias = [f"Recepcion orden de compra #{oid}" for oid in orden_ids]
    movimientos_propios = _kardex_por_referencia_in(db, referencias_propias)
    lote_ids = {m.lote_id for m in movimientos_propios}
    posteriores = _kardex_posteriores(db, lote_ids, referencias_propias=referencias_propias)
    if posteriores:
        referencias = sorted({m.referencia or "(sin referencia)" for m in posteriores})
        bloqueos.append(schemas.BloqueoReemplazoOut(
            tipo="MOVIMIENTOS_DERIVADOS",
            detalle=(
                f"{len(posteriores)} movimiento(s) de kardex posterior(es) (ventas/ajustes ya "
                "aplicados sobre la mercaderia de estas compras) generados por: " + "; ".join(referencias)
            ),
        ))

    # "Pagos" (Cuentas por Pagar): el ERP todavia no tiene ese modulo, por
    # lo que esta categoria no aplica hoy (no hay tabla que verificar).
    # Se deja documentado para cuando exista.

    return schemas.ValidacionReemplazoOut(
        carga_id=carga.id,
        tipo_carga="COMPRAS",
        estado_carga=carga.estado,
        estado_vigencia=carga.estado_vigencia,
        puede_reemplazar=len(bloqueos) == 0,
        bloqueos=bloqueos,
    )


def reemplazar_compras(
    db: Session,
    carga_id: int,
    nombre_archivo: str,
    contenido: bytes,
    motivo: str,
    observaciones: str | None,
    usuario,
    ip: str | None = None,
) -> schemas.ReemplazoOut:
    inicio = time.monotonic()
    validators.validar_permiso_reemplazo(usuario)
    motivo = validators.validar_motivo_obligatorio(motivo)

    carga = obtener_carga_compras(db, carga_id)
    validators.validar_carga_confirmada(carga)
    validators.validar_carga_no_reemplazada(carga)

    validacion = verificar_reemplazo_compras(db, carga_id)
    if not validacion.puede_reemplazar:
        tiempo_ms = int((time.monotonic() - inicio) * 1000)
        _bitacora_error(
            db, "COMPRAS", carga.id, usuario, motivo, observaciones, ip,
            "BLOQUEADO", "; ".join(f"[{b.tipo}] {b.detalle}" for b in validacion.bloqueos), tiempo_ms,
        )
    validators.validar_sin_bloqueos(validacion.bloqueos)

    try:
        preview = previsualizar_compras(db, carga.inventario_id, nombre_archivo, contenido)
    except ValueError as exc:
        tiempo_ms = int((time.monotonic() - inicio) * 1000)
        _bitacora_error(db, "COMPRAS", carga.id, usuario, motivo, observaciones, ip, "ERROR", str(exc), tiempo_ms)
        raise

    if preview.filas_con_error:
        detalle = (
            f"El archivo de reemplazo tiene {preview.filas_con_error} fila(s) con error "
            f"(carga de previsualizacion #{preview.carga_id}, sin confirmar). "
            "Corrija el Excel y vuelva a intentar el reemplazo; la carga original no fue modificada."
        )
        tiempo_ms = int((time.monotonic() - inicio) * 1000)
        _bitacora_error(db, "COMPRAS", carga.id, usuario, motivo, observaciones, ip, "ERROR", detalle, tiempo_ms)
        raise ValueError(detalle)

    try:
        with _transaccion_atomica(db):
            filas = repository.filas_procesadas_compras(db, carga_id)
            orden_ids = list({f.orden_compra_id for f in filas if f.orden_compra_id is not None})
            referencias_propias = [f"Recepcion orden de compra #{oid}" for oid in orden_ids]
            movimientos = _kardex_por_referencia_in(db, referencias_propias)
            lote_ids = list({m.lote_id for m in movimientos})
            kardex_eliminados = len(movimientos)
            for m in movimientos:
                db.delete(m)
            db.flush()

            lotes_eliminados = 0
            if lote_ids:
                lotes = db.query(Lote).filter(Lote.id.in_(lote_ids)).all()
                lotes_eliminados = len(lotes)
                for l in lotes:
                    db.delete(l)
            db.flush()

            # Desvincular las filas de la carga vieja de la orden que se va
            # a borrar (evita referencias huerfanas a orden_compra_id).
            ordenes_eliminadas = 0
            if orden_ids:
                for fila in filas:
                    if fila.orden_compra_id in orden_ids:
                        fila.orden_compra_id = None
                        db.add(fila)
                db.flush()
                ordenes = db.query(OrdenCompra).filter(OrdenCompra.id.in_(orden_ids)).all()
                ordenes_eliminadas = len(ordenes)
                for o in ordenes:
                    db.delete(o)  # cascade="all, delete-orphan" borra sus items
            db.flush()

            resultado_confirmacion = confirmar_compras(db, preview.carga_id)
            if resultado_confirmacion.filas_fallidas_en_confirmacion:
                raise ValueError(
                    "La carga de reemplazo no se pudo confirmar por completo "
                    f"({len(resultado_confirmacion.filas_fallidas_en_confirmacion)} fila(s) fallaron "
                    "pese a haber pasado la previsualizacion). Se revirtio todo el reemplazo, "
                    "la carga original sigue vigente sin cambios."
                )

            nueva_carga = obtener_carga_compras(db, preview.carga_id)
            repository.marcar_carga_reemplazada_compras(db, carga, nueva_carga, getattr(usuario, "id", None), motivo, observaciones)

            tiempo_ms = int((time.monotonic() - inicio) * 1000)
            bitacora = BitacoraReemplazo(
                tipo_carga="COMPRAS",
                carga_anterior_id=carga.id,
                carga_nueva_id=nueva_carga.id,
                usuario_id=getattr(usuario, "id", None),
                usuario_username=getattr(usuario, "username", "desconocido"),
                ip_origen=ip,
                motivo=motivo,
                observaciones=observaciones,
                cantidad_lotes_eliminados=lotes_eliminados,
                cantidad_kardex_eliminados=kardex_eliminados,
                cantidad_ordenes_eliminadas=ordenes_eliminadas,
                cantidad_registros_nuevos=resultado_confirmacion.filas_procesadas,
                tiempo_ejecucion_ms=tiempo_ms,
                resultado="EXITOSO",
                detalle=None,
            )
            db.add(bitacora)
            db.flush()
            db.refresh(bitacora)
    except Exception as exc:
        tiempo_ms = int((time.monotonic() - inicio) * 1000)
        _bitacora_error(db, "COMPRAS", carga.id, usuario, motivo, observaciones, ip, "ERROR", str(exc), tiempo_ms)
        raise

    db.refresh(carga)
    return schemas.ReemplazoOut(
        carga_anterior_id=carga.id,
        carga_anterior_estado_vigencia=carga.estado_vigencia,
        carga_nueva_id=nueva_carga.id,
        carga_nueva_estado_vigencia=nueva_carga.estado_vigencia,
        kardex_eliminados=kardex_eliminados,
        lotes_eliminados=lotes_eliminados,
        ordenes_eliminadas=ordenes_eliminadas,
        filas_procesadas=resultado_confirmacion.filas_procesadas,
        filas_fallidas_en_confirmacion=resultado_confirmacion.filas_fallidas_en_confirmacion,
        motivo=motivo,
        tiempo_ejecucion_ms=tiempo_ms,
        bitacora_id=bitacora.id,
    )


# -----------------------------------------------------------------
# VENTAS HISTORICO
# -----------------------------------------------------------------


def verificar_reemplazo_ventas(db: Session, carga_id: int) -> schemas.ValidacionReemplazoOut:
    carga = obtener_carga_ventas(db, carga_id)
    bloqueos: list[schemas.BloqueoReemplazoOut] = []

    if carga.estado != "CONFIRMADA":
        bloqueos.append(schemas.BloqueoReemplazoOut(
            tipo="ESTADO",
            detalle=f"La carga esta en estado '{carga.estado}', debe estar CONFIRMADA para poder reemplazarse.",
        ))
    if carga.estado_vigencia == "REEMPLAZADA":
        bloqueos.append(schemas.BloqueoReemplazoOut(
            tipo="YA_REEMPLAZADA",
            detalle=f"Esta carga ya fue reemplazada por la carga #{carga.carga_reemplazo_id}.",
        ))

    filas = repository.filas_procesadas_ventas(db, carga_id)
    orden_ids = sorted({f.orden_venta_id for f in filas if f.orden_venta_id is not None})

    if orden_ids:
        guias = db.query(GuiaRemision).filter(GuiaRemision.orden_venta_id.in_(orden_ids)).all()
        if guias:
            ordenes_afectadas = sorted({g.orden_venta_id for g in guias})
            numeros = sorted({g.numero_guia for g in guias})
            bloqueos.append(schemas.BloqueoReemplazoOut(
                tipo="GUIA_REMISION",
                detalle=(
                    f"{len(guias)} guia(s) de remision emitida(s) ({', '.join(numeros)}) "
                    f"sobre la(s) orden(es) de venta {ordenes_afectadas}."
                ),
            ))

        operaciones = db.query(OperacionLogistica).filter(OperacionLogistica.orden_venta_id.in_(orden_ids)).all()
        if operaciones:
            ordenes_afectadas = sorted({o.orden_venta_id for o in operaciones})
            bloqueos.append(schemas.BloqueoReemplazoOut(
                tipo="OPERACION_LOGISTICA",
                detalle=(
                    f"{len(operaciones)} operacion(es) logistica(s) (picking/packing/carga/despacho) "
                    f"registradas sobre la(s) orden(es) de venta {ordenes_afectadas}."
                ),
            ))

    referencias_propias = [f"Despacho orden de venta #{oid}" for oid in orden_ids]
    movimientos_propios = _kardex_por_referencia_in(db, referencias_propias)
    lote_ids = {m.lote_id for m in movimientos_propios}
    posteriores = _kardex_posteriores(db, lote_ids, referencias_propias=referencias_propias)
    if posteriores:
        referencias = sorted({m.referencia or "(sin referencia)" for m in posteriores})
        bloqueos.append(schemas.BloqueoReemplazoOut(
            tipo="MOVIMIENTOS_DERIVADOS",
            detalle=(
                f"{len(posteriores)} movimiento(s) de kardex posterior(es) generados por: " + "; ".join(referencias)
            ),
        ))

    # "Facturas" (facturacion electronica propia): el ERP no tiene ese
    # modulo hoy (m12_sunat no define un modelo Factura); no aplica,
    # documentado para cuando exista.

    return schemas.ValidacionReemplazoOut(
        carga_id=carga.id,
        tipo_carga="VENTAS",
        estado_carga=carga.estado,
        estado_vigencia=carga.estado_vigencia,
        puede_reemplazar=len(bloqueos) == 0,
        bloqueos=bloqueos,
    )


def reemplazar_ventas(
    db: Session,
    carga_id: int,
    nombre_archivo: str,
    contenido: bytes,
    motivo: str,
    observaciones: str | None,
    usuario,
    ip: str | None = None,
) -> schemas.ReemplazoOut:
    inicio = time.monotonic()
    validators.validar_permiso_reemplazo(usuario)
    motivo = validators.validar_motivo_obligatorio(motivo)

    carga = obtener_carga_ventas(db, carga_id)
    validators.validar_carga_confirmada(carga)
    validators.validar_carga_no_reemplazada(carga)

    validacion = verificar_reemplazo_ventas(db, carga_id)
    if not validacion.puede_reemplazar:
        tiempo_ms = int((time.monotonic() - inicio) * 1000)
        _bitacora_error(
            db, "VENTAS", carga.id, usuario, motivo, observaciones, ip,
            "BLOQUEADO", "; ".join(f"[{b.tipo}] {b.detalle}" for b in validacion.bloqueos), tiempo_ms,
        )
    validators.validar_sin_bloqueos(validacion.bloqueos)

    try:
        preview = previsualizar_ventas(db, carga.inventario_id, nombre_archivo, contenido)
    except ValueError as exc:
        tiempo_ms = int((time.monotonic() - inicio) * 1000)
        _bitacora_error(db, "VENTAS", carga.id, usuario, motivo, observaciones, ip, "ERROR", str(exc), tiempo_ms)
        raise

    if preview.filas_con_error:
        detalle = (
            f"El archivo de reemplazo tiene {preview.filas_con_error} fila(s) con error "
            f"(carga de previsualizacion #{preview.carga_id}, sin confirmar). "
            "Corrija el Excel y vuelva a intentar el reemplazo; la carga original no fue modificada."
        )
        tiempo_ms = int((time.monotonic() - inicio) * 1000)
        _bitacora_error(db, "VENTAS", carga.id, usuario, motivo, observaciones, ip, "ERROR", detalle, tiempo_ms)
        raise ValueError(detalle)

    try:
        with _transaccion_atomica(db):
            filas = repository.filas_procesadas_ventas(db, carga_id)
            orden_ids = list({f.orden_venta_id for f in filas if f.orden_venta_id is not None})
            referencias_propias = [f"Despacho orden de venta #{oid}" for oid in orden_ids]
            movimientos = _kardex_por_referencia_in(db, referencias_propias)
            lote_ids = list({m.lote_id for m in movimientos})
            kardex_eliminados = len(movimientos)
            for m in movimientos:
                db.delete(m)
            db.flush()

            lotes_eliminados = 0
            if lote_ids:
                lotes = db.query(Lote).filter(Lote.id.in_(lote_ids)).all()
                lotes_eliminados = len(lotes)
                for l in lotes:
                    db.delete(l)
            db.flush()

            ordenes_eliminadas = 0
            if orden_ids:
                for fila in filas:
                    if fila.orden_venta_id in orden_ids:
                        fila.orden_venta_id = None
                        db.add(fila)
                db.flush()
                ordenes = db.query(OrdenVenta).filter(OrdenVenta.id.in_(orden_ids)).all()
                ordenes_eliminadas = len(ordenes)
                for o in ordenes:
                    db.delete(o)
            db.flush()

            resultado_confirmacion = confirmar_ventas(db, preview.carga_id)
            if resultado_confirmacion.filas_fallidas_en_confirmacion:
                raise ValueError(
                    "La carga de reemplazo no se pudo confirmar por completo "
                    f"({len(resultado_confirmacion.filas_fallidas_en_confirmacion)} fila(s) fallaron "
                    "pese a haber pasado la previsualizacion). Se revirtio todo el reemplazo, "
                    "la carga original sigue vigente sin cambios."
                )

            nueva_carga = obtener_carga_ventas(db, preview.carga_id)
            repository.marcar_carga_reemplazada_ventas(db, carga, nueva_carga, getattr(usuario, "id", None), motivo, observaciones)

            tiempo_ms = int((time.monotonic() - inicio) * 1000)
            bitacora = BitacoraReemplazo(
                tipo_carga="VENTAS",
                carga_anterior_id=carga.id,
                carga_nueva_id=nueva_carga.id,
                usuario_id=getattr(usuario, "id", None),
                usuario_username=getattr(usuario, "username", "desconocido"),
                ip_origen=ip,
                motivo=motivo,
                observaciones=observaciones,
                cantidad_lotes_eliminados=lotes_eliminados,
                cantidad_kardex_eliminados=kardex_eliminados,
                cantidad_ordenes_eliminadas=ordenes_eliminadas,
                cantidad_registros_nuevos=resultado_confirmacion.filas_procesadas,
                tiempo_ejecucion_ms=tiempo_ms,
                resultado="EXITOSO",
                detalle=None,
            )
            db.add(bitacora)
            db.flush()
            db.refresh(bitacora)
    except Exception as exc:
        tiempo_ms = int((time.monotonic() - inicio) * 1000)
        _bitacora_error(db, "VENTAS", carga.id, usuario, motivo, observaciones, ip, "ERROR", str(exc), tiempo_ms)
        raise

    db.refresh(carga)
    return schemas.ReemplazoOut(
        carga_anterior_id=carga.id,
        carga_anterior_estado_vigencia=carga.estado_vigencia,
        carga_nueva_id=nueva_carga.id,
        carga_nueva_estado_vigencia=nueva_carga.estado_vigencia,
        kardex_eliminados=kardex_eliminados,
        lotes_eliminados=lotes_eliminados,
        ordenes_eliminadas=ordenes_eliminadas,
        filas_procesadas=resultado_confirmacion.filas_procesadas,
        filas_fallidas_en_confirmacion=resultado_confirmacion.filas_fallidas_en_confirmacion,
        motivo=motivo,
        tiempo_ejecucion_ms=tiempo_ms,
        bitacora_id=bitacora.id,
    )
